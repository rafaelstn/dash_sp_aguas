"""
Worker de indexação do HD de rede SPÁguas (v1.4 — Modelo A + Modelo B).

Estratégia (ver docs/architecture.md §8 e docs/repositorio-cliente.md §3):

  Modelo A — pastas por prefixo (Fluviometria/, Pluviometria/, Piezometria/,
             QualiAgua/, e aninhados Fluviometria/ANA, Fluviometria/QualiAgua):
    1. Determina `tipo_dado` pela posição do arquivo na hierarquia.
    2. Trata pastas especiais (case-insensitive, sem acento):
         "0 SEM PREFIXO" / "0 SemPrefixo"       -> órfão PREFIXO_DESCONHECIDO
         "0 DUVIDAS"     / "0 duvidas"          -> órfão PENDENCIA_CLIENTE
         "0 Paralisados" / "0 Paralisado"       -> indexa + flag paralisado
         "0 FICHAS DESCRITIVAS *"               -> órfão FICHA_GERAL
         "{PREFIXO} paralisado" (ex.: "1D-008 paralisado") -> indexa + flag
    3. Normaliza whitespace e parseia o nome contra 3 regex (em ordem de
       especificidade): COMPLETO -> PARCIAL -> LEGADO (ver v1.3).
    4. Valida prefixo capturado contra a regex do tipo_dado e faz lookup:
         - FluviometriaANA: LPAD(prefixo_ana, 8, '0')
         - demais: igualdade em postos.prefixo
    5. Buckets: arquivos_indexados (match total) / arquivos_orfaos
       (PREFIXO_DESCONHECIDO se parse OK mas sem lookup / NOME_FORA_DO_PADRAO
       se nenhum dos 3 formatos bateu).

  Modelo B — REPOSITORIO CTHDOC (nome = número sequencial):
    1. Arquivos sob `DOCUMENTOS-CTHDOC-RECUPERAÇÃO/REPOSITORIO/{numero}.pdf`.
    2. Vínculo com posto vem da planilha `relacao_doc_arquivos_cthdoc.xlsx`:
         col A = PREFIXO, col B = Tipo Dado, col C = Tipo Documento, col F = Arquivo.
    3. Traduz texto da planilha para códigos via sinônimos normalizados.
    4. Buckets: arquivos_indexados (match na planilha + lookup do prefixo em
       postos) / arquivos_orfaos PREFIXO_DESCONHECIDO (planilha não cita ou
       prefixo inexistente).
    5. Grava com origem_mapeamento='PLANILHA_XLSX', formato_nome='REPOSITORIO',
       numero_arquivo={numero}.

Uso:
  python index_fs.py [--root PATH] [--dry-run] [--mapping-xlsx PATH]

Variáveis de ambiente:
  DATABASE_URL_INDEXER   conn string com permissão de escrita nas tabelas alvo
  INDEXER_ROOT_PATH      raiz se --root não for passado

--dry-run: varre e classifica tudo sem abrir transação de escrita; imprime
sumário por bucket + amostras. Use antes de qualquer execução real.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Iterable
from uuid import uuid4

import psycopg
import structlog
from dotenv import load_dotenv
from unidecode import unidecode

log = structlog.get_logger(__name__)

EXTENSOES_ACEITAS = {".pdf"}

# ---------------------------------------------------------------------------
# Regex de nome de arquivo — Modelo A
# ---------------------------------------------------------------------------

RX_WHITESPACE = re.compile(r"\s{2,}")

# COMPLETO — padrão oficial novo (apenas 7,88% da base histórica):
#   {Prefixo} {CodDoc:2d} {CodEnc:2d} [opcional] {AAAA MM DD}.pdf
RX_COMPLETO = re.compile(
    r"^(?P<prefixo>\S+)\s+(?P<cod_doc>\d{2})\s+(?P<cod_enc>\d{2})"
    r"(?:\s+(?P<opcional>.+?))?\s+(?P<data>\d{4}\s+\d{2}\s+\d{2})\.pdf$",
    flags=re.IGNORECASE,
)

# PARCIAL — prefixo + cod_doc + data:
#   {Prefixo} {CodDoc:2d} {AAAA MM DD}.pdf  (ex.: "1D-002 01 1960 08 26.pdf")
RX_PARCIAL = re.compile(
    r"^(?P<prefixo>\S+)\s+(?P<cod_doc>\d{2})\s+(?P<data>\d{4}\s+\d{2}\s+\d{2})\.pdf$",
    flags=re.IGNORECASE,
)

# LEGADO — apenas prefixo + data:
#   {Prefixo} {AAAA MM DD}.pdf
RX_LEGADO = re.compile(
    r"^(?P<prefixo>\S+)\s+(?P<data>\d{4}\s+\d{2}\s+\d{2})\.pdf$",
    flags=re.IGNORECASE,
)

# Modelo B — número puro (com ou sem zero à esquerda):
RX_REPOSITORIO_NUMERO = re.compile(r"^(?P<numero>\d+)\.pdf$", flags=re.IGNORECASE)

# Pasta "{prefixo} {sufixo_textual}" (ex.: "1D-008 paralisado"):
RX_PASTA_COM_SUFIXO = re.compile(r"^(?P<prefixo>\S+)\s+(?P<sufixo>.+)$")

# ---------------------------------------------------------------------------
# Constantes de pastas especiais
# ---------------------------------------------------------------------------

PASTAS_SEM_PREFIXO = {"0 sem prefixo", "0 semprefixo"}
PASTAS_DUVIDAS = {"0 duvidas"}
PASTAS_PARALISADAS = {"0 paralisados", "0 paralisado"}
# Prefixos de pastas tratadas como FICHA_GERAL (documentos institucionais, não por posto).
#   "0 fichas descritivas *"  — fichas gerais do programa (Fluviometria/Pluviometria).
#   "0 boletins"              — boletins institucionais anuais (Fluviometria/0 Boletins/
#                                2019..2024/OUTROS); confirmado via mapeamento real do HD.
PREFIXOS_PASTA_FICHA_GERAL = ("0 fichas descritivas", "0 boletins")

XLSX_CTHDOC_NOME = "relacao_doc_arquivos_cthdoc.xlsx"

# Fragmento presente em qualquer pasta raiz do Modelo B (tolera variações de acento).
RX_PASTA_CTHDOC = re.compile(r"documentos.cthdoc.recuper", flags=re.IGNORECASE)
PASTA_REPOSITORIO = "repositorio"

# ---------------------------------------------------------------------------
# Dicionários de tradução planilha XLSX -> códigos do banco
# ---------------------------------------------------------------------------

# Texto da coluna B (Tipo Dado) -> código em `tipos_dado`.
SINONIMOS_TIPO_DADO: dict[str, str] = {
    "PLUVIAL": "Pluviometria",
    "PLUVIOMETRIA": "Pluviometria",
    "PLUVIOMETRICO": "Pluviometria",
    "FLUVIAL": "Fluviometria",
    "FLUVIOMETRIA": "Fluviometria",
    "FLUVIOMETRICO": "Fluviometria",
    "FLUVIAL ANA": "FluviometriaANA",
    "FLUVIOMETRIA ANA": "FluviometriaANA",
    "ANA": "FluviometriaANA",
    "FLUVIAL QUALIAGUA": "FluviometriaQualiAgua",
    "FLUVIOMETRIA QUALIAGUA": "FluviometriaQualiAgua",
    "PIEZOMETRIA": "Piezometria",
    "PIEZOMETRICO": "Piezometria",
    "QUALIAGUA": "QualiAgua",
    "QUALI AGUA": "QualiAgua",
}

# Texto da coluna C (Tipo Documento) -> código em `tipos_documento` (01..07).
SINONIMOS_COD_TIPO_DOC: dict[str, int] = {
    "FICHA DESCRITIVA": 1,
    "DESCRITIVA": 1,
    "PCD": 2,
    "FICHA DE INSPECAO": 3,
    "INSPECAO": 3,
    "FICHA DE INSPECAO DE CAMPO": 3,
    "NIVELAMENTO": 4,
    "LEVANTAMENTO DE SECAO": 5,
    "LEVANTAMENTO DE SECCAO": 5,
    "LEVANTAMENTO": 5,
    "TROCA DE OBSERVADOR": 6,
    "OBSERVADOR": 6,
    "VAZAO": 7,
    "MEDICAO DE VAZAO": 7,
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(texto: str) -> str:
    """Minúsculas, sem acento, whitespace colapsado."""
    return re.sub(r"\s+", " ", unidecode(texto).strip().lower())


def _norm_upper(texto: str) -> str:
    return re.sub(r"\s+", " ", unidecode(texto).strip().upper())


def _normalizar_nome(nome: str) -> str:
    """Colapsa whitespace múltiplo em espaço único."""
    return RX_WHITESPACE.sub(" ", nome).strip()


def _parse_nome_modelo_a(nome: str) -> tuple[str, dict] | None:
    """Bate o nome contra COMPLETO -> PARCIAL -> LEGADO. None se nenhum bater."""
    m = RX_COMPLETO.match(nome)
    if m is not None:
        return "COMPLETO", {
            "prefixo": m.group("prefixo"),
            "cod_doc": m.group("cod_doc"),
            "cod_enc": m.group("cod_enc"),
            "opcional": m.group("opcional"),
            "data": m.group("data"),
        }

    m = RX_PARCIAL.match(nome)
    if m is not None:
        return "PARCIAL", {
            "prefixo": m.group("prefixo"),
            "cod_doc": m.group("cod_doc"),
            "cod_enc": None,
            "opcional": None,
            "data": m.group("data"),
        }

    m = RX_LEGADO.match(nome)
    if m is not None:
        return "LEGADO", {
            "prefixo": m.group("prefixo"),
            "cod_doc": None,
            "cod_enc": None,
            "opcional": None,
            "data": m.group("data"),
        }

    return None


def _parse_data(texto: str) -> date | None:
    partes = texto.split()
    if len(partes) != 3:
        return None
    try:
        ano, mes, dia = int(partes[0]), int(partes[1]), int(partes[2])
        return date(ano, mes, dia)
    except ValueError:
        return None


def _partes_relativas(caminho: Path, raiz: Path) -> list[str]:
    """Partes da hierarquia (pastas entre raiz e arquivo, excluindo o nome)."""
    try:
        rel = caminho.relative_to(raiz)
    except ValueError:
        return []
    return list(rel.parts[:-1])


def _detectar_modelo(caminho: Path, raiz: Path) -> str:
    """'B' se o arquivo está sob DOCUMENTOS-CTHDOC-RECUPER*/REPOSITORIO; senão 'A'."""
    partes = [_norm(p) for p in _partes_relativas(caminho, raiz)]
    for i, parte in enumerate(partes):
        if RX_PASTA_CTHDOC.search(parte):
            # A pasta REPOSITORIO precisa existir como descendente imediato.
            if PASTA_REPOSITORIO in partes[i + 1 :]:
                return "B"
    return "A"


def _detectar_tipo_dado(caminho: Path, raiz: Path) -> str | None:
    """Determina tipo_dado pela posição na hierarquia (Modelo A).

    Considera tanto as partes RELATIVAS (subpastas abaixo da raiz) quanto as
    partes DA RAIZ — isso permite apontar --root pra uma subpasta (ex.:
    Y:\\Fluviometria\\1D-008\\) sem perder o contexto de TipoDado. Útil para
    cargas seletivas.
    """
    partes_raiz = [_norm(p) for p in raiz.parts]
    partes_rel = [_norm(p) for p in _partes_relativas(caminho, raiz)]
    partes = partes_raiz + partes_rel
    if not partes:
        return None

    for i, parte in enumerate(partes):
        if parte == "fluviometria":
            if i + 1 < len(partes):
                sub = partes[i + 1]
                if sub == "ana":
                    return "FluviometriaANA"
                if sub == "qualiagua":
                    return "FluviometriaQualiAgua"
            return "Fluviometria"
        if parte == "pluviometria":
            return "Pluviometria"
        if parte == "piezometria":
            return "Piezometria"
        if parte == "qualiagua":
            return "QualiAgua"
    return None


def _classificar_pasta_especial(
    caminho: Path, raiz: Path
) -> tuple[str | None, bool]:
    """Retorna (categoria_orfao, flag_paralisado).

    categoria_orfao ∈ {None, 'PREFIXO_DESCONHECIDO', 'PENDENCIA_CLIENTE', 'FICHA_GERAL'}
    flag_paralisado = True se algum ancestral é pasta de paralisados ou
    '{prefixo} paralisado'.
    """
    partes = _partes_relativas(caminho, raiz)
    categoria: str | None = None
    paralisado = False

    for parte in partes:
        norm = _norm(parte)
        if norm in PASTAS_SEM_PREFIXO:
            categoria = "PREFIXO_DESCONHECIDO"
        elif norm in PASTAS_DUVIDAS:
            categoria = "PENDENCIA_CLIENTE"
        elif norm in PASTAS_PARALISADAS:
            paralisado = True
        elif any(norm.startswith(p) for p in PREFIXOS_PASTA_FICHA_GERAL):
            categoria = "FICHA_GERAL"
        else:
            m = RX_PASTA_COM_SUFIXO.match(parte)
            if m and "paralisa" in _norm(m.group("sufixo")):
                paralisado = True

    return categoria, paralisado


def _marcar_paralisado(opcional: str | None) -> str:
    """Concatena a flag textual '[paralisado]' ao campo parte_opcional."""
    tag = "[paralisado]"
    if opcional:
        return f"{tag} {opcional}"
    return tag


# ---------------------------------------------------------------------------
# Catálogos do banco
# ---------------------------------------------------------------------------

def _carregar_tipos_dado(conn: psycopg.Connection) -> dict[str, dict]:
    with conn.cursor() as cur:
        cur.execute("SELECT codigo, regex_prefixo, usa_prefixo_ana FROM tipos_dado")
        return {
            row[0]: {"regex": re.compile(row[1]), "usa_prefixo_ana": row[2]}
            for row in cur.fetchall()
        }


def _carregar_tipos_documento(conn: psycopg.Connection) -> dict[str, int]:
    """Retorna {rotulo_normalizado -> codigo}."""
    with conn.cursor() as cur:
        cur.execute("SELECT codigo, rotulo FROM tipos_documento")
        return {_norm_upper(row[1]): int(row[0]) for row in cur.fetchall()}


def _carregar_prefixos(conn: psycopg.Connection) -> set[str]:
    with conn.cursor() as cur:
        cur.execute("SELECT prefixo FROM postos")
        return {row[0] for row in cur.fetchall()}


def _carregar_prefixos_ana(conn: psycopg.Connection) -> dict[str, str]:
    """{prefixo_ana padded(8) -> prefixo do posto}."""
    with conn.cursor() as cur:
        cur.execute(
            "SELECT prefixo, prefixo_ana FROM postos "
            "WHERE prefixo_ana IS NOT NULL AND prefixo_ana <> ''"
        )
        mapa: dict[str, str] = {}
        for prefixo, prefixo_ana in cur.fetchall():
            chave = prefixo_ana.zfill(8) if prefixo_ana.isdigit() else prefixo_ana
            mapa[chave] = prefixo
        return mapa


# ---------------------------------------------------------------------------
# Mapa CTHDOC (planilha XLSX)
# ---------------------------------------------------------------------------

def _mapear_tipo_dado_texto(texto: str, codigos_validos: set[str]) -> str | None:
    if not texto:
        return None
    chave = _norm_upper(texto).replace("-", " ")
    chave = re.sub(r"\s+", " ", chave)
    candidato = SINONIMOS_TIPO_DADO.get(chave)
    if candidato and candidato in codigos_validos:
        return candidato
    # Fallback: bate com o próprio código (unaccent+upper).
    for codigo in codigos_validos:
        if _norm_upper(codigo) == chave:
            return codigo
    return None


def _mapear_cod_tipo_doc_texto(
    texto: str, rotulos_normalizados: dict[str, int]
) -> int | None:
    if not texto:
        return None
    chave = _norm_upper(texto)
    if chave in SINONIMOS_COD_TIPO_DOC:
        return SINONIMOS_COD_TIPO_DOC[chave]
    if chave in rotulos_normalizados:
        return rotulos_normalizados[chave]
    return None


def _localizar_xlsx(raiz: Path, override: Path | None) -> Path | None:
    if override:
        return override if override.exists() else None
    candidatos = [
        raiz / "DOCUMENTOS-CTHDOC-RECUPERAÇÃO" / XLSX_CTHDOC_NOME,
        raiz / "DOCUMENTOS-CTHDOC-RECUPERACAO" / XLSX_CTHDOC_NOME,
        raiz / XLSX_CTHDOC_NOME,
    ]
    for cand in candidatos:
        if cand.exists():
            return cand
    # Busca rasa (até 3 níveis) como último recurso.
    for cand in raiz.glob(f"*/{XLSX_CTHDOC_NOME}"):
        return cand
    for cand in raiz.glob(f"**/{XLSX_CTHDOC_NOME}"):
        return cand
    return None


def _carregar_mapa_cthdoc(
    xlsx: Path,
    codigos_tipo_dado: set[str],
    rotulos_tipo_doc: dict[str, int],
) -> dict[str, dict]:
    """Lê a planilha e retorna {numero_sem_ext: {prefixo, tipo_dado, cod_tipo_documento}}.

    Colunas (manual do cliente): A=PREFIXO, B=Tipo Dado, C=Tipo Documento, F=Arquivo.
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl_nao_instalado", xlsx=str(xlsx))
        return {}

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active

    mapa: dict[str, dict] = {}
    primeira_linha = True
    for row in ws.iter_rows(values_only=True):
        if primeira_linha:
            primeira_linha = False
            continue
        if not row or len(row) < 6:
            continue

        prefixo_raw = row[0]
        tipo_dado_txt = row[1]
        tipo_doc_txt = row[2]
        arquivo_raw = row[5]

        if not prefixo_raw or not arquivo_raw:
            continue

        prefixo = str(prefixo_raw).strip()
        arquivo_nome = str(arquivo_raw).strip()
        if not prefixo or not arquivo_nome:
            continue

        numero = Path(arquivo_nome).stem
        if not numero:
            continue

        tipo_dado = _mapear_tipo_dado_texto(
            str(tipo_dado_txt) if tipo_dado_txt else "",
            codigos_tipo_dado,
        )
        cod_doc = _mapear_cod_tipo_doc_texto(
            str(tipo_doc_txt) if tipo_doc_txt else "",
            rotulos_tipo_doc,
        )

        mapa[numero] = {
            "prefixo": prefixo,
            "tipo_dado": tipo_dado,
            "cod_tipo_documento": cod_doc,
        }

    wb.close()
    return mapa


# ---------------------------------------------------------------------------
# Varredura
# ---------------------------------------------------------------------------

def _varrer(raiz: Path) -> Iterable[Path]:
    for caminho in raiz.rglob("*"):
        if not caminho.is_file():
            continue
        if caminho.suffix.lower() not in EXTENSOES_ACEITAS:
            continue
        yield caminho


# ---------------------------------------------------------------------------
# SQL de persistência
# ---------------------------------------------------------------------------

SQL_UPSERT_INDEXADO = """
INSERT INTO arquivos_indexados
  (prefixo, nome_arquivo, caminho_absoluto, tamanho_bytes,
   data_modificacao, lote_indexacao,
   tipo_dado, cod_tipo_documento, cod_encarregado,
   data_documento, parte_opcional, nome_valido, formato_nome,
   numero_arquivo, origem_mapeamento)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, true, %s, %s, %s)
ON CONFLICT (caminho_absoluto) DO UPDATE SET
  prefixo              = EXCLUDED.prefixo,
  nome_arquivo         = EXCLUDED.nome_arquivo,
  tamanho_bytes        = EXCLUDED.tamanho_bytes,
  data_modificacao     = EXCLUDED.data_modificacao,
  lote_indexacao       = EXCLUDED.lote_indexacao,
  tipo_dado            = EXCLUDED.tipo_dado,
  cod_tipo_documento   = EXCLUDED.cod_tipo_documento,
  cod_encarregado      = EXCLUDED.cod_encarregado,
  data_documento       = EXCLUDED.data_documento,
  parte_opcional       = EXCLUDED.parte_opcional,
  nome_valido          = true,
  formato_nome         = EXCLUDED.formato_nome,
  numero_arquivo       = EXCLUDED.numero_arquivo,
  origem_mapeamento    = EXCLUDED.origem_mapeamento,
  indexado_em          = NOW()
"""

SQL_UPSERT_ORFAO = """
INSERT INTO arquivos_orfaos
  (nome_arquivo, caminho_absoluto, tamanho_bytes, data_modificacao,
   lote_indexacao, categoria, tipo_dado, motivo)
VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
ON CONFLICT (caminho_absoluto) DO UPDATE SET
  nome_arquivo     = EXCLUDED.nome_arquivo,
  tamanho_bytes    = EXCLUDED.tamanho_bytes,
  data_modificacao = EXCLUDED.data_modificacao,
  lote_indexacao   = EXCLUDED.lote_indexacao,
  categoria        = EXCLUDED.categoria,
  tipo_dado        = EXCLUDED.tipo_dado,
  motivo           = EXCLUDED.motivo,
  indexado_em      = NOW()
"""

MOTIVO_POR_CATEGORIA = {
    "NOME_FORA_DO_PADRAO": "nome_fora_do_padrao",
    "PREFIXO_DESCONHECIDO": "prefixo_nao_identificado",
    "PENDENCIA_CLIENTE": "pendencia_cliente",
    "FICHA_GERAL": "ficha_geral_sem_prefixo",
}


# ---------------------------------------------------------------------------
# Classificação por arquivo
# ---------------------------------------------------------------------------

def _classificar_arquivo(
    arquivo: Path,
    raiz: Path,
    *,
    tipos_dado: dict[str, dict],
    prefixos_conhecidos: set[str],
    prefixos_ana: dict[str, str],
    mapa_cthdoc: dict[str, dict] | None,
) -> dict:
    """Retorna dict descrevendo onde o arquivo vai (indexado/órfão) + payload.

    Não faz IO de banco. Pode ser chamado em dry-run e no modo real.
    """
    modelo = _detectar_modelo(arquivo, raiz)

    if modelo == "B":
        return _classificar_modelo_b(arquivo, raiz, mapa_cthdoc, tipos_dado)

    return _classificar_modelo_a(
        arquivo,
        raiz,
        tipos_dado=tipos_dado,
        prefixos_conhecidos=prefixos_conhecidos,
        prefixos_ana=prefixos_ana,
    )


def _classificar_modelo_a(
    arquivo: Path,
    raiz: Path,
    *,
    tipos_dado: dict[str, dict],
    prefixos_conhecidos: set[str],
    prefixos_ana: dict[str, str],
) -> dict:
    tipo_dado = _detectar_tipo_dado(arquivo, raiz)

    categoria_especial, paralisado = _classificar_pasta_especial(arquivo, raiz)
    if categoria_especial is not None:
        return {
            "bucket": "orfao",
            "categoria": categoria_especial,
            "tipo_dado": tipo_dado,
        }

    nome_norm = _normalizar_nome(arquivo.name)
    parsed = _parse_nome_modelo_a(nome_norm)

    if parsed is None:
        return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}

    formato_nome, grupos = parsed
    prefixo_capturado = grupos["prefixo"]
    cod_doc_raw = grupos["cod_doc"]
    cod_enc = grupos["cod_enc"]
    opcional = grupos["opcional"]
    data_documento = _parse_data(grupos["data"])

    if tipo_dado is None or tipo_dado not in tipos_dado:
        return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}

    cfg_tipo = tipos_dado[tipo_dado]
    if not cfg_tipo["regex"].match(prefixo_capturado):
        return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}

    cod_doc: int | None = None
    if cod_doc_raw is not None:
        try:
            cod_doc_int = int(cod_doc_raw)
        except ValueError:
            cod_doc_int = 0
        if cod_doc_int < 1 or cod_doc_int > 7:
            return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}
        cod_doc = cod_doc_int

    if cfg_tipo["usa_prefixo_ana"]:
        chave = (
            prefixo_capturado.zfill(8)
            if prefixo_capturado.isdigit()
            else prefixo_capturado
        )
        prefixo_final = prefixos_ana.get(chave)
    else:
        prefixo_final = (
            prefixo_capturado if prefixo_capturado in prefixos_conhecidos else None
        )

    if prefixo_final is None:
        return {"bucket": "orfao", "categoria": "PREFIXO_DESCONHECIDO", "tipo_dado": tipo_dado}

    opcional_final = _marcar_paralisado(opcional) if paralisado else opcional

    return {
        "bucket": "indexado",
        "prefixo": prefixo_final,
        "tipo_dado": tipo_dado,
        "cod_tipo_documento": cod_doc,
        "cod_encarregado": cod_enc,
        "data_documento": data_documento,
        "parte_opcional": opcional_final,
        "formato_nome": formato_nome,
        "numero_arquivo": None,
        "origem_mapeamento": "NOME",
    }


def _classificar_modelo_b(
    arquivo: Path,
    raiz: Path,
    mapa_cthdoc: dict[str, dict] | None,
    tipos_dado: dict[str, dict],
) -> dict:
    tipo_dado_fallback: str | None = None

    m = RX_REPOSITORIO_NUMERO.match(arquivo.name)
    if m is None:
        return {
            "bucket": "orfao",
            "categoria": "NOME_FORA_DO_PADRAO",
            "tipo_dado": tipo_dado_fallback,
        }

    numero = m.group("numero")

    # Sem planilha: todos viram órfão PREFIXO_DESCONHECIDO.
    if mapa_cthdoc is None:
        return {
            "bucket": "orfao",
            "categoria": "PREFIXO_DESCONHECIDO",
            "tipo_dado": tipo_dado_fallback,
        }

    entrada = mapa_cthdoc.get(numero) or mapa_cthdoc.get(numero.lstrip("0"))
    if entrada is None:
        return {
            "bucket": "orfao",
            "categoria": "PREFIXO_DESCONHECIDO",
            "tipo_dado": tipo_dado_fallback,
        }

    prefixo = entrada["prefixo"]
    tipo_dado = entrada.get("tipo_dado") or tipo_dado_fallback
    cod_doc = entrada.get("cod_tipo_documento")

    # Sanidade: se tipo_dado vier None ou não estiver na tabela, órfão.
    if tipo_dado is None or tipo_dado not in tipos_dado:
        return {
            "bucket": "orfao",
            "categoria": "PREFIXO_DESCONHECIDO",
            "tipo_dado": None,
        }

    return {
        "bucket": "indexado",
        "prefixo": prefixo,
        "tipo_dado": tipo_dado,
        "cod_tipo_documento": cod_doc,
        "cod_encarregado": None,
        "data_documento": None,
        "parte_opcional": None,
        "formato_nome": "REPOSITORIO",
        "numero_arquivo": numero,
        "origem_mapeamento": "PLANILHA_XLSX",
    }


# ---------------------------------------------------------------------------
# Execução
# ---------------------------------------------------------------------------

def executar(
    raiz: Path,
    conn_str: str,
    *,
    dry_run: bool,
    mapping_xlsx: Path | None,
) -> int:
    if not raiz.exists() or not raiz.is_dir():
        log.error("raiz_invalida", raiz=str(raiz))
        return 3

    lote = str(uuid4())
    log.info(
        "lote_iniciado",
        lote=lote,
        raiz=str(raiz),
        modo="dry_run" if dry_run else "persistencia",
    )

    with psycopg.connect(conn_str) as conn:
        tipos_dado = _carregar_tipos_dado(conn)
        rotulos_tipo_doc = _carregar_tipos_documento(conn)
        prefixos_conhecidos = _carregar_prefixos(conn)
        prefixos_ana = _carregar_prefixos_ana(conn)
        log.info(
            "catalogos_carregados",
            tipos_dado=len(tipos_dado),
            tipos_documento=len(rotulos_tipo_doc),
            prefixos=len(prefixos_conhecidos),
            prefixos_ana=len(prefixos_ana),
        )

        mapa_cthdoc: dict[str, dict] | None = None
        xlsx_path = _localizar_xlsx(raiz, mapping_xlsx)
        if xlsx_path:
            mapa_cthdoc = _carregar_mapa_cthdoc(
                xlsx_path,
                set(tipos_dado.keys()),
                rotulos_tipo_doc,
            )
            log.info(
                "mapa_cthdoc_carregado",
                xlsx=str(xlsx_path),
                entradas=len(mapa_cthdoc),
            )
        else:
            log.warning("mapa_cthdoc_nao_encontrado", xlsx_nome=XLSX_CTHDOC_NOME)

        sumario = {
            "total": 0,
            "indexado_modelo_a": 0,
            "indexado_modelo_b": 0,
            "orfao_prefixo_desconhecido": 0,
            "orfao_nome_fora_padrao": 0,
            "orfao_pendencia_cliente": 0,
            "orfao_ficha_geral": 0,
            "paralisados_detectados": 0,
        }
        amostras: dict[str, list[str]] = {k: [] for k in sumario.keys()}
        erros: list[dict] = []

        with conn.cursor() as cur:
            if not dry_run:
                cur.execute(
                    "INSERT INTO indexacao_log (lote_indexacao, raiz_varredura, status) "
                    "VALUES (%s, %s, 'em_andamento')",
                    (lote, str(raiz)),
                )
                conn.commit()

            for arquivo in _varrer(raiz):
                sumario["total"] += 1
                try:
                    stat = arquivo.stat()
                    mtime = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc)

                    resultado = _classificar_arquivo(
                        arquivo,
                        raiz,
                        tipos_dado=tipos_dado,
                        prefixos_conhecidos=prefixos_conhecidos,
                        prefixos_ana=prefixos_ana,
                        mapa_cthdoc=mapa_cthdoc,
                    )

                    if resultado["bucket"] == "indexado":
                        chave = (
                            "indexado_modelo_b"
                            if resultado["origem_mapeamento"] == "PLANILHA_XLSX"
                            else "indexado_modelo_a"
                        )
                        sumario[chave] += 1
                        if len(amostras[chave]) < 10:
                            amostras[chave].append(str(arquivo))
                        opcional_txt = resultado.get("parte_opcional") or ""
                        if opcional_txt.startswith("[paralisado]"):
                            sumario["paralisados_detectados"] += 1

                        if not dry_run:
                            cur.execute(
                                SQL_UPSERT_INDEXADO,
                                (
                                    resultado["prefixo"],
                                    arquivo.name,
                                    str(arquivo),
                                    stat.st_size,
                                    mtime,
                                    lote,
                                    resultado["tipo_dado"],
                                    resultado["cod_tipo_documento"],
                                    resultado["cod_encarregado"],
                                    resultado["data_documento"],
                                    resultado["parte_opcional"],
                                    resultado["formato_nome"],
                                    resultado["numero_arquivo"],
                                    resultado["origem_mapeamento"],
                                ),
                            )
                    else:
                        categoria = resultado["categoria"]
                        chave = {
                            "PREFIXO_DESCONHECIDO": "orfao_prefixo_desconhecido",
                            "NOME_FORA_DO_PADRAO": "orfao_nome_fora_padrao",
                            "PENDENCIA_CLIENTE": "orfao_pendencia_cliente",
                            "FICHA_GERAL": "orfao_ficha_geral",
                        }[categoria]
                        sumario[chave] += 1
                        if len(amostras[chave]) < 10:
                            amostras[chave].append(str(arquivo))

                        if not dry_run:
                            cur.execute(
                                SQL_UPSERT_ORFAO,
                                (
                                    arquivo.name,
                                    str(arquivo),
                                    stat.st_size,
                                    mtime,
                                    lote,
                                    categoria,
                                    resultado["tipo_dado"],
                                    MOTIVO_POR_CATEGORIA[categoria],
                                ),
                            )
                except (OSError, psycopg.Error) as e:
                    erros.append({"caminho": str(arquivo), "erro": str(e)})
                    if not dry_run:
                        conn.rollback()

            if not dry_run:
                conn.commit()

                prefixo_raiz = str(raiz)
                cur.execute(
                    "DELETE FROM arquivos_indexados WHERE lote_indexacao <> %s "
                    "AND caminho_absoluto LIKE %s",
                    (lote, f"{prefixo_raiz}%"),
                )
                removidos_indexados = cur.rowcount
                cur.execute(
                    "DELETE FROM arquivos_orfaos WHERE lote_indexacao <> %s "
                    "AND caminho_absoluto LIKE %s",
                    (lote, f"{prefixo_raiz}%"),
                )
                removidos_orfaos = cur.rowcount

                total_orfaos = (
                    sumario["orfao_prefixo_desconhecido"]
                    + sumario["orfao_nome_fora_padrao"]
                    + sumario["orfao_pendencia_cliente"]
                    + sumario["orfao_ficha_geral"]
                )
                total_indexados = (
                    sumario["indexado_modelo_a"] + sumario["indexado_modelo_b"]
                )

                cur.execute(
                    "UPDATE indexacao_log SET finalizado_em = NOW(), "
                    "arquivos_encontrados = %s, arquivos_indexados_qtd = %s, "
                    "arquivos_orfaos_qtd = %s, arquivos_removidos_qtd = %s, "
                    "erros_amostra = %s::jsonb, status = 'ok' "
                    "WHERE lote_indexacao = %s",
                    (
                        sumario["total"],
                        total_indexados,
                        total_orfaos,
                        removidos_indexados + removidos_orfaos,
                        json.dumps(erros[:20], ensure_ascii=False),
                        lote,
                    ),
                )
                conn.commit()

    relatorio = {
        "lote": lote,
        "raiz": str(raiz),
        "modo": "dry_run" if dry_run else "persistencia",
        "xlsx_mapeamento": str(xlsx_path) if xlsx_path else None,
        "sumario": sumario,
        "amostras_por_bucket": amostras,
        "erros_qtd": len(erros),
    }

    print(json.dumps(relatorio, ensure_ascii=False, indent=2, default=str))

    log.info(
        "lote_concluido",
        lote=lote,
        total=sumario["total"],
        indexado_a=sumario["indexado_modelo_a"],
        indexado_b=sumario["indexado_modelo_b"],
        orfaos=sum(v for k, v in sumario.items() if k.startswith("orfao_")),
        erros=len(erros),
        modo=relatorio["modo"],
    )
    return 0


def main() -> int:
    load_dotenv(".env.local")
    load_dotenv(".env")

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--root", default=os.environ.get("INDEXER_ROOT_PATH"))
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Varre e classifica sem gravar no banco. Imprime sumário JSON.",
    )
    parser.add_argument(
        "--mapping-xlsx",
        default=None,
        help="Caminho da planilha relacao_doc_arquivos_cthdoc.xlsx (override do auto-detect).",
    )
    args = parser.parse_args()

    raiz = args.root
    if not raiz:
        print(
            "erro: informe --root ou defina INDEXER_ROOT_PATH no .env.local",
            file=sys.stderr,
        )
        return 1

    conn_str = os.environ.get("DATABASE_URL_INDEXER") or os.environ.get("DATABASE_URL")
    if not conn_str:
        print(
            "erro: defina DATABASE_URL_INDEXER (ou DATABASE_URL) no .env.local",
            file=sys.stderr,
        )
        return 1

    mapping = Path(args.mapping_xlsx) if args.mapping_xlsx else None
    return executar(Path(raiz), conn_str, dry_run=args.dry_run, mapping_xlsx=mapping)


if __name__ == "__main__":
    raise SystemExit(main())
