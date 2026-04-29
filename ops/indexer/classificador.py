"""
Engine de classificação do indexador SPÁguas (extraído de index_fs.py em 2026-04).

Este módulo concentra tudo que é PURE (sem IO de filesystem além de Path.name
e sem IO de banco além de SELECTs idempotentes de catálogo):

  - Regex Modelo A / B.
  - Parser de nome.
  - Detecção de tipo_dado pela hierarquia.
  - Classificação de pastas especiais (sem prefixo, duvidas, paralisados,
    fichas gerais, "{prefixo} paralisado").
  - Carga de catálogos (tipos_dado, tipos_documento, prefixos, prefixos_ana).
  - Carga do mapa CTHDOC (planilha XLSX).
  - Função-raiz `classificar_arquivo` que retorna o dict padronizado de
    bucket/categoria/payload que o persistidor usa.

Regra de ouro: NENHUMA função aqui escreve em banco. Quem persiste é
`index_fs.py` (sweep completo) ou `indexar_posto.py` (on-demand por posto).

Compat: `index_fs.py` reexporta os nomes legados `_parse_nome_modelo_a`,
`_detectar_tipo_dado`, `_classificar_pasta_especial`, `_carregar_tipos_dado`,
`_carregar_prefixos`, `_carregar_prefixos_ana`, `_carregar_mapa_cthdoc`,
`_classificar_arquivo`.
"""

from __future__ import annotations

import re
from datetime import date
from pathlib import Path

import psycopg
import structlog
from unidecode import unidecode

log = structlog.get_logger(__name__)

# ---------------------------------------------------------------------------
# Constantes públicas
# ---------------------------------------------------------------------------

EXTENSOES_ACEITAS = {".pdf"}

RX_WHITESPACE = re.compile(r"\s{2,}")

RX_COMPLETO = re.compile(
    r"^(?P<prefixo>\S+)\s+(?P<cod_doc>\d{2})\s+(?P<cod_enc>\d{2})"
    r"(?:\s+(?P<opcional>.+?))?\s+(?P<data>\d{4}\s+\d{2}\s+\d{2})\.pdf$",
    flags=re.IGNORECASE,
)

RX_PARCIAL = re.compile(
    r"^(?P<prefixo>\S+)\s+(?P<cod_doc>\d{2})\s+(?P<data>\d{4}\s+\d{2}\s+\d{2})\.pdf$",
    flags=re.IGNORECASE,
)

RX_LEGADO = re.compile(
    r"^(?P<prefixo>\S+)\s+(?P<data>\d{4}\s+\d{2}\s+\d{2})\.pdf$",
    flags=re.IGNORECASE,
)

RX_REPOSITORIO_NUMERO = re.compile(r"^(?P<numero>\d+)\.pdf$", flags=re.IGNORECASE)
RX_PASTA_COM_SUFIXO = re.compile(r"^(?P<prefixo>\S+)\s+(?P<sufixo>.+)$")

PASTAS_SEM_PREFIXO = {"0 sem prefixo", "0 semprefixo"}
PASTAS_DUVIDAS = {"0 duvidas"}
PASTAS_PARALISADAS = {"0 paralisados", "0 paralisado"}
PREFIXOS_PASTA_FICHA_GERAL = ("0 fichas descritivas", "0 boletins")

XLSX_CTHDOC_NOME = "relacao_doc_arquivos_cthdoc.xlsx"
RX_PASTA_CTHDOC = re.compile(r"documentos.cthdoc.recuper", flags=re.IGNORECASE)
PASTA_REPOSITORIO = "repositorio"

SINONIMOS_TIPO_DADO: dict[str, str] = {
    "PLUVIAL": "Pluviometria",
    "PLUVIOMETRIA": "Pluviometria",
    "PLUVIOMETRICO": "Pluviometria",
    "FLUVIAL": "Fluviometria",
    "FLUVIOMETRIA": "Fluviometria",
    "FLUVIOMETRICO": "Fluviometria",
    "FLUVIAL ANA": "FluviometriaANA",
    "FLUVIOMETRIA ANA": "FluviometriaANA",
    "ANA": "FluviometriaANA",
    "FLUVIAL QUALIAGUA": "FluviometriaQualiAgua",
    "FLUVIOMETRIA QUALIAGUA": "FluviometriaQualiAgua",
    "PIEZOMETRIA": "Piezometria",
    "PIEZOMETRICO": "Piezometria",
    "QUALIAGUA": "QualiAgua",
    "QUALI AGUA": "QualiAgua",
}

SINONIMOS_COD_TIPO_DOC: dict[str, int] = {
    "FICHA DESCRITIVA": 1,
    "DESCRITIVA": 1,
    "PCD": 2,
    "FICHA DE INSPECAO": 3,
    "INSPECAO": 3,
    "FICHA DE INSPECAO DE CAMPO": 3,
    "NIVELAMENTO": 4,
    "LEVANTAMENTO DE SECAO": 5,
    "LEVANTAMENTO DE SECCAO": 5,
    "LEVANTAMENTO": 5,
    "TROCA DE OBSERVADOR": 6,
    "OBSERVADOR": 6,
    "VAZAO": 7,
    "MEDICAO DE VAZAO": 7,
}


# ---------------------------------------------------------------------------
# Helpers de string
# ---------------------------------------------------------------------------

def _norm(texto: str) -> str:
    return re.sub(r"\s+", " ", unidecode(texto).strip().lower())


def _norm_upper(texto: str) -> str:
    return re.sub(r"\s+", " ", unidecode(texto).strip().upper())


def _normalizar_nome(nome: str) -> str:
    return RX_WHITESPACE.sub(" ", nome).strip()


# ---------------------------------------------------------------------------
# Parser de nome de arquivo
# ---------------------------------------------------------------------------

def parse_nome_modelo_a(nome: str) -> tuple[str, dict] | None:
    """Bate o nome contra COMPLETO -> PARCIAL -> LEGADO. None se nenhum bater."""
    m = RX_COMPLETO.match(nome)
    if m is not None:
        return "COMPLETO", {
            "prefixo": m.group("prefixo"),
            "cod_doc": m.group("cod_doc"),
            "cod_enc": m.group("cod_enc"),
            "opcional": m.group("opcional"),
            "data": m.group("data"),
        }

    m = RX_PARCIAL.match(nome)
    if m is not None:
        return "PARCIAL", {
            "prefixo": m.group("prefixo"),
            "cod_doc": m.group("cod_doc"),
            "cod_enc": None,
            "opcional": None,
            "data": m.group("data"),
        }

    m = RX_LEGADO.match(nome)
    if m is not None:
        return "LEGADO", {
            "prefixo": m.group("prefixo"),
            "cod_doc": None,
            "cod_enc": None,
            "opcional": None,
            "data": m.group("data"),
        }

    return None


def parse_data(texto: str) -> date | None:
    partes = texto.split()
    if len(partes) != 3:
        return None
    try:
        ano, mes, dia = int(partes[0]), int(partes[1]), int(partes[2])
        return date(ano, mes, dia)
    except ValueError:
        return None


# ---------------------------------------------------------------------------
# Detecção hierárquica
# ---------------------------------------------------------------------------

def partes_relativas(caminho: Path, raiz: Path) -> list[str]:
    try:
        rel = caminho.relative_to(raiz)
    except ValueError:
        return []
    return list(rel.parts[:-1])


def detectar_modelo(caminho: Path, raiz: Path) -> str:
    """'B' se o arquivo está sob DOCUMENTOS-CTHDOC-RECUPER*/REPOSITORIO; senão 'A'."""
    partes = [_norm(p) for p in partes_relativas(caminho, raiz)]
    for i, parte in enumerate(partes):
        if RX_PASTA_CTHDOC.search(parte):
            if PASTA_REPOSITORIO in partes[i + 1 :]:
                return "B"
    return "A"


def detectar_tipo_dado(caminho: Path, raiz: Path) -> str | None:
    """Determina tipo_dado pela posição na hierarquia (Modelo A).

    Considera partes da RAIZ + partes RELATIVAS para suportar --root apontando
    pra subpasta (ex.: Y:\\000 Documentos de Campo\\Fluviometria\\1D-008\\).
    """
    partes_raiz = [_norm(p) for p in raiz.parts]
    partes_rel = [_norm(p) for p in partes_relativas(caminho, raiz)]
    partes = partes_raiz + partes_rel
    if not partes:
        return None

    for i, parte in enumerate(partes):
        if parte == "fluviometria":
            if i + 1 < len(partes):
                sub = partes[i + 1]
                if sub == "ana":
                    return "FluviometriaANA"
                if sub == "qualiagua":
                    return "FluviometriaQualiAgua"
            return "Fluviometria"
        if parte == "pluviometria":
            return "Pluviometria"
        if parte == "piezometria":
            return "Piezometria"
        if parte == "qualiagua":
            return "QualiAgua"
    return None


def classificar_pasta_especial(
    caminho: Path, raiz: Path
) -> tuple[str | None, bool]:
    """Retorna (categoria_orfao, flag_paralisado).

    categoria_orfao ∈ {None, 'PREFIXO_DESCONHECIDO', 'PENDENCIA_CLIENTE', 'FICHA_GERAL'}
    """
    partes = partes_relativas(caminho, raiz)
    categoria: str | None = None
    paralisado = False

    for parte in partes:
        norm = _norm(parte)
        if norm in PASTAS_SEM_PREFIXO:
            categoria = "PREFIXO_DESCONHECIDO"
        elif norm in PASTAS_DUVIDAS:
            categoria = "PENDENCIA_CLIENTE"
        elif norm in PASTAS_PARALISADAS:
            paralisado = True
        elif any(norm.startswith(p) for p in PREFIXOS_PASTA_FICHA_GERAL):
            categoria = "FICHA_GERAL"
        else:
            m = RX_PASTA_COM_SUFIXO.match(parte)
            if m and "paralisa" in _norm(m.group("sufixo")):
                paralisado = True

    return categoria, paralisado


def marcar_paralisado(opcional: str | None) -> str:
    tag = "[paralisado]"
    if opcional:
        return f"{tag} {opcional}"
    return tag


# ---------------------------------------------------------------------------
# Catálogos (IO de banco — SELECT idempotente, sem escrita)
# ---------------------------------------------------------------------------

def carregar_tipos_dado(conn: psycopg.Connection) -> dict[str, dict]:
    with conn.cursor() as cur:
        cur.execute("SELECT codigo, regex_prefixo, usa_prefixo_ana FROM tipos_dado")
        return {
            row[0]: {"regex": re.compile(row[1]), "usa_prefixo_ana": row[2]}
            for row in cur.fetchall()
        }


def carregar_tipos_documento(conn: psycopg.Connection) -> dict[str, int]:
    with conn.cursor() as cur:
        cur.execute("SELECT codigo, rotulo FROM tipos_documento")
        return {_norm_upper(row[1]): int(row[0]) for row in cur.fetchall()}


def carregar_prefixos(conn: psycopg.Connection) -> set[str]:
    with conn.cursor() as cur:
        cur.execute("SELECT prefixo FROM postos")
        return {row[0] for row in cur.fetchall()}


def carregar_prefixos_ana(conn: psycopg.Connection) -> dict[str, str]:
    with conn.cursor() as cur:
        cur.execute(
            "SELECT prefixo, prefixo_ana FROM postos "
            "WHERE prefixo_ana IS NOT NULL AND prefixo_ana <> ''"
        )
        mapa: dict[str, str] = {}
        for prefixo, prefixo_ana in cur.fetchall():
            chave = prefixo_ana.zfill(8) if prefixo_ana.isdigit() else prefixo_ana
            mapa[chave] = prefixo
        return mapa


# ---------------------------------------------------------------------------
# Mapa CTHDOC
# ---------------------------------------------------------------------------

def _mapear_tipo_dado_texto(texto: str, codigos_validos: set[str]) -> str | None:
    if not texto:
        return None
    chave = _norm_upper(texto).replace("-", " ")
    chave = re.sub(r"\s+", " ", chave)
    candidato = SINONIMOS_TIPO_DADO.get(chave)
    if candidato and candidato in codigos_validos:
        return candidato
    for codigo in codigos_validos:
        if _norm_upper(codigo) == chave:
            return codigo
    return None


def _mapear_cod_tipo_doc_texto(
    texto: str, rotulos_normalizados: dict[str, int]
) -> int | None:
    if not texto:
        return None
    chave = _norm_upper(texto)
    if chave in SINONIMOS_COD_TIPO_DOC:
        return SINONIMOS_COD_TIPO_DOC[chave]
    if chave in rotulos_normalizados:
        return rotulos_normalizados[chave]
    return None


def localizar_xlsx(raiz: Path, override: Path | None) -> Path | None:
    if override:
        return override if override.exists() else None
    candidatos = [
        raiz / "DOCUMENTOS-CTHDOC-RECUPERAÇÃO" / XLSX_CTHDOC_NOME,
        raiz / "DOCUMENTOS-CTHDOC-RECUPERACAO" / XLSX_CTHDOC_NOME,
        raiz / XLSX_CTHDOC_NOME,
    ]
    for cand in candidatos:
        if cand.exists():
            return cand
    for cand in raiz.glob(f"*/{XLSX_CTHDOC_NOME}"):
        return cand
    for cand in raiz.glob(f"**/{XLSX_CTHDOC_NOME}"):
        return cand
    return None


def carregar_mapa_cthdoc(
    xlsx: Path,
    codigos_tipo_dado: set[str],
    rotulos_tipo_doc: dict[str, int],
) -> dict[str, dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        log.error("openpyxl_nao_instalado", xlsx=str(xlsx))
        return {}

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    ws = wb.active

    mapa: dict[str, dict] = {}
    primeira_linha = True
    for row in ws.iter_rows(values_only=True):
        if primeira_linha:
            primeira_linha = False
            continue
        if not row or len(row) < 6:
            continue

        prefixo_raw = row[0]
        tipo_dado_txt = row[1]
        tipo_doc_txt = row[2]
        arquivo_raw = row[5]

        if not prefixo_raw or not arquivo_raw:
            continue

        prefixo = str(prefixo_raw).strip()
        arquivo_nome = str(arquivo_raw).strip()
        if not prefixo or not arquivo_nome:
            continue

        numero = Path(arquivo_nome).stem
        if not numero:
            continue

        tipo_dado = _mapear_tipo_dado_texto(
            str(tipo_dado_txt) if tipo_dado_txt else "",
            codigos_tipo_dado,
        )
        cod_doc = _mapear_cod_tipo_doc_texto(
            str(tipo_doc_txt) if tipo_doc_txt else "",
            rotulos_tipo_doc,
        )

        mapa[numero] = {
            "prefixo": prefixo,
            "tipo_dado": tipo_dado,
            "cod_tipo_documento": cod_doc,
        }

    wb.close()
    return mapa


# ---------------------------------------------------------------------------
# Classificação por arquivo — ENTRYPOINT
# ---------------------------------------------------------------------------

def classificar_arquivo(
    arquivo: Path,
    raiz: Path,
    *,
    tipos_dado: dict[str, dict],
    prefixos_conhecidos: set[str],
    prefixos_ana: dict[str, str],
    mapa_cthdoc: dict[str, dict] | None,
) -> dict:
    """Retorna dict descrevendo o destino do arquivo (indexado/órfão) + payload.

    Não faz IO de banco. Pode ser chamado em dry-run e no modo real.
    """
    modelo = detectar_modelo(arquivo, raiz)

    if modelo == "B":
        return _classificar_modelo_b(arquivo, raiz, mapa_cthdoc, tipos_dado)

    return _classificar_modelo_a(
        arquivo,
        raiz,
        tipos_dado=tipos_dado,
        prefixos_conhecidos=prefixos_conhecidos,
        prefixos_ana=prefixos_ana,
    )


def _classificar_modelo_a(
    arquivo: Path,
    raiz: Path,
    *,
    tipos_dado: dict[str, dict],
    prefixos_conhecidos: set[str],
    prefixos_ana: dict[str, str],
) -> dict:
    tipo_dado = detectar_tipo_dado(arquivo, raiz)

    categoria_especial, paralisado = classificar_pasta_especial(arquivo, raiz)
    if categoria_especial is not None:
        return {
            "bucket": "orfao",
            "categoria": categoria_especial,
            "tipo_dado": tipo_dado,
        }

    nome_norm = _normalizar_nome(arquivo.name)
    parsed = parse_nome_modelo_a(nome_norm)

    if parsed is None:
        return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}

    formato_nome, grupos = parsed
    prefixo_capturado = grupos["prefixo"]
    cod_doc_raw = grupos["cod_doc"]
    cod_enc = grupos["cod_enc"]
    opcional = grupos["opcional"]
    data_documento = parse_data(grupos["data"])

    if tipo_dado is None or tipo_dado not in tipos_dado:
        return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}

    cfg_tipo = tipos_dado[tipo_dado]
    if not cfg_tipo["regex"].match(prefixo_capturado):
        return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}

    cod_doc: int | None = None
    if cod_doc_raw is not None:
        try:
            cod_doc_int = int(cod_doc_raw)
        except ValueError:
            cod_doc_int = 0
        if cod_doc_int < 1 or cod_doc_int > 7:
            return {"bucket": "orfao", "categoria": "NOME_FORA_DO_PADRAO", "tipo_dado": tipo_dado}
        cod_doc = cod_doc_int

    if cfg_tipo["usa_prefixo_ana"]:
        chave = (
            prefixo_capturado.zfill(8)
            if prefixo_capturado.isdigit()
            else prefixo_capturado
        )
        prefixo_final = prefixos_ana.get(chave)
    else:
        prefixo_final = (
            prefixo_capturado if prefixo_capturado in prefixos_conhecidos else None
        )

    if prefixo_final is None:
        return {"bucket": "orfao", "categoria": "PREFIXO_DESCONHECIDO", "tipo_dado": tipo_dado}

    opcional_final = marcar_paralisado(opcional) if paralisado else opcional

    return {
        "bucket": "indexado",
        "prefixo": prefixo_final,
        "tipo_dado": tipo_dado,
        "cod_tipo_documento": cod_doc,
        "cod_encarregado": cod_enc,
        "data_documento": data_documento,
        "parte_opcional": opcional_final,
        "formato_nome": formato_nome,
        "numero_arquivo": None,
        "origem_mapeamento": "NOME",
    }


def _classificar_modelo_b(
    arquivo: Path,
    raiz: Path,
    mapa_cthdoc: dict[str, dict] | None,
    tipos_dado: dict[str, dict],
) -> dict:
    tipo_dado_fallback: str | None = None

    m = RX_REPOSITORIO_NUMERO.match(arquivo.name)
    if m is None:
        return {
            "bucket": "orfao",
            "categoria": "NOME_FORA_DO_PADRAO",
            "tipo_dado": tipo_dado_fallback,
        }

    numero = m.group("numero")

    if mapa_cthdoc is None:
        return {
            "bucket": "orfao",
            "categoria": "PREFIXO_DESCONHECIDO",
            "tipo_dado": tipo_dado_fallback,
        }

    entrada = mapa_cthdoc.get(numero) or mapa_cthdoc.get(numero.lstrip("0"))
    if entrada is None:
        return {
            "bucket": "orfao",
            "categoria": "PREFIXO_DESCONHECIDO",
            "tipo_dado": tipo_dado_fallback,
        }

    prefixo = entrada["prefixo"]
    tipo_dado = entrada.get("tipo_dado") or tipo_dado_fallback
    cod_doc = entrada.get("cod_tipo_documento")

    if tipo_dado is None or tipo_dado not in tipos_dado:
        return {
            "bucket": "orfao",
            "categoria": "PREFIXO_DESCONHECIDO",
            "tipo_dado": None,
        }

    return {
        "bucket": "indexado",
        "prefixo": prefixo,
        "tipo_dado": tipo_dado,
        "cod_tipo_documento": cod_doc,
        "cod_encarregado": None,
        "data_documento": None,
        "parte_opcional": None,
        "formato_nome": "REPOSITORIO",
        "numero_arquivo": numero,
        "origem_mapeamento": "PLANILHA_XLSX",
    }
