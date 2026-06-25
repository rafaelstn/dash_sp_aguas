"""Aplica as correções SPÁguas direto na planilha original do SharePoint da ANA.

Lê a planilha "PROGESTÃO 3 META I.6 ..." baixada do SharePoint (inventário
nacional com ~11.7k linhas) e, somente nas linhas com Responsável-UF = "SP"
(~2371 linhas), substitui os valores das colunas pelos valores efetivos do
banco SPÁguas (fallback: postos > resposta_* > snapshot ANA).

Para cada célula alterada:
  - escreve o novo valor
  - pinta o fundo de amarelo (#FFFF00) conforme exigência da ANA

Adiciona 2 colunas ao final da aba Inventário (38 e 39):
  - STATUS_REVISAO_SPAGUAS
  - JUSTIFICATIVA_SPAGUAS

Demais abas (Instruções, Responsavel_Preenchimento, VersãoPlanilha) e
linhas de outros estados ficam intactas.

Saída: <pasta-da-origem>/<basename-da-origem>_RESPOSTA_SPAGUAS.xlsx
Idempotente: re-execução sobrescreve a saída usando o estado atual do banco.

Uso:
  ops/indexer/.venv/Scripts/python.exe scripts/manutencao/aplicar_resposta_na_planilha_sharepoint.py \\
      --origem "C:/Users/win1064/Downloads/PROGESTÃO_3...xlsx"
  (opcional) --saida "C:/Users/win1064/Downloads/saida.xlsx"
  (opcional) --dry-run  (não salva, só relata estatísticas)
"""
from __future__ import annotations

import argparse
import json
import os
import unicodedata
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path

import psycopg
from dotenv import load_dotenv
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

ROOT = Path(__file__).resolve().parents[2]
load_dotenv(ROOT / ".env.local")
DB_URL = os.environ["DATABASE_URL"]

# Schema das colunas vem do arquivo unico compartilhado com o export TS
# (data/colunas-ana.json). Editar la muda os dois lados.
SCHEMA = json.loads((ROOT / "data" / "colunas-ana.json").read_text(encoding="utf-8"))

AMARELO = PatternFill(fill_type="solid", fgColor=SCHEMA["corDiffArgb"])

TIPO_TEXTO = "texto"
TIPO_NUMERO = "numero"
TIPO_DATA = "data"
TIPO_BOOL = "bool"

# Mapeamento coluna (1-based) -> (alias no SELECT, tipo, label), derivado do
# schema compartilhado. A query de buscar_estado() devolve um alias por chave.
COLUNAS = [
    (c["colExcel"], c["aliasPy"], c["tipo"], c["label"])
    for c in SCHEMA["colunas"]
]

LABEL_STATUS = SCHEMA["colunasControle"]["status"]["label"]
LABEL_JUSTIFICATIVA = SCHEMA["colunasControle"]["justificativa"]["label"]


def chave_nome_municipio(nome: str) -> str:
    """Normaliza nome de municipio para lookup case e acento insensitivo:
    minusculas + remove diacriticos. Mesma regra do export TS
    (chaveNomeMunicipio em exportar.ts)."""
    base = unicodedata.normalize("NFD", nome.lower())
    sem_acento = "".join(c for c in base if unicodedata.category(c) != "Mn")
    return sem_acento.strip()


def carregar_mapa_ibge(cur) -> dict[str, str]:
    """nome (normalizado) -> codigo_ibge, a partir de ibge_municipios_sp.

    Sem este lookup, o XLSX pode sair com nome de um municipio e codigo IBGE
    de outro (Q08 da auditoria 2026-05-18). Espelha exportar.ts.
    """
    cur.execute("SELECT codigo_ibge, nome FROM ibge_municipios_sp")
    mapa: dict[str, str] = {}
    for codigo_ibge, nome in cur.fetchall():
        mapa[chave_nome_municipio(str(nome))] = str(codigo_ibge)
    return mapa


def buscar_estado() -> dict[str, dict]:
    """Carrega o estado efetivo de todas as estações SP em memoria.

    Retorna dict[codigo_ana] -> dict com os campos efetivos.
    """
    sql_query = """
    SELECT
      e.codigo_ana,
      -- nome
      COALESCE(p.nome_estacao, e.nome)                                AS nome_efetivo,
      -- codigo_adicional (prefixo DAEE)
      COALESCE(p.prefixo, e.codigo_adicional)                         AS codigo_adicional_efetivo,
      -- coord
      COALESCE(p.latitude::text, e.resposta_latitude::text, e.latitude::text) AS latitude_efetiva,
      COALESCE(p.longitude::text, e.resposta_longitude::text, e.longitude::text) AS longitude_efetiva,
      -- demais
      COALESCE(p.altimetria::text, e.altitude::text)                  AS altitude_efetiva,
      COALESCE(p.area_km2::text, e.area_drenagem_km2::text)           AS area_efetiva,
      COALESCE(p.bacia_hidrografica, e.bacia_nome)                    AS bacia_nome_efetiva,
      COALESCE(p.sub_ugrhi_nome, e.subbacia_nome)                     AS subbacia_nome_efetiva,
      -- codigo do municipio: fallback resposta_* > ANA. Quando o nome
      -- efetivo vem de `postos`, o codigo IBGE eh resolvido no Python pelo
      -- nome (ver pos-processamento), espelhando exportar.ts (Q08).
      COALESCE(e.resposta_municipio_codigo, e.municipio_codigo)       AS municipio_codigo_efetivo,
      COALESCE(p.municipio, e.resposta_municipio_nome, e.municipio_nome) AS municipio_nome_efetivo,
      -- brutos usados so para decidir a regra do codigo do municipio
      p.municipio                                                     AS _p_municipio,
      e.municipio_nome                                                AS _ana_municipio_nome,
      COALESCE(p.tipo_posto, e.estacao_tipo)                          AS estacao_tipo_efetiva,
      COALESCE(p.ana_escala_inicio, e.escala_inicio)                  AS escala_inicio_efetiva,
      COALESCE(p.ana_escala_fim, e.escala_fim)                        AS escala_fim_efetiva,
      COALESCE(p.ana_descarga_liquida_inicio, e.descarga_liquida_inicio) AS descarga_inicio_efetiva,
      COALESCE(p.ana_descarga_liquida_fim, e.descarga_liquida_fim)    AS descarga_fim_efetiva,
      COALESCE(p.ana_sedimentos_inicio, e.sedimentos_inicio)          AS sedimentos_inicio_efetiva,
      COALESCE(p.ana_sedimentos_fim, e.sedimentos_fim)                AS sedimentos_fim_efetiva,
      COALESCE(p.ana_qualidade_inicio, e.qualidade_inicio)            AS qualidade_inicio_efetiva,
      COALESCE(p.ana_qualidade_fim, e.qualidade_fim)                  AS qualidade_fim_efetiva,
      COALESCE(p.ana_pluviometro_inicio, e.pluviometro_inicio)        AS pluviometro_inicio_efetiva,
      COALESCE(p.ana_pluviometro_fim, e.pluviometro_fim)              AS pluviometro_fim_efetiva,
      COALESCE(p.ana_telemetria_inicio, e.telemetria_inicio)          AS telemetria_inicio_efetiva,
      COALESCE(p.ana_telemetria_fim, e.telemetria_fim)                AS telemetria_fim_efetiva,
      e.status,
      e.resposta_justificativa
      FROM ana_revisao_estacao e
      LEFT JOIN postos p ON p.id = e.posto_id AND p.deleted_at IS NULL
     WHERE e.lote_id = (SELECT id FROM ana_revisao_lote ORDER BY criado_em DESC LIMIT 1)
    """

    estado: dict[str, dict] = {}
    with psycopg.connect(DB_URL, prepare_threshold=None) as conn, conn.cursor() as cur:
        mapa_ibge = carregar_mapa_ibge(cur)
        cur.execute(sql_query)
        cols = [d[0] for d in cur.description]
        for row in cur.fetchall():
            r = dict(zip(cols, row))
            # Q08: quando o nome efetivo veio de `postos` e difere do nome ANA,
            # o codigo do municipio precisa vir do IBGE pelo nome efetivo,
            # senao o XLSX sai com "Atibaia, cod de outro municipio".
            p_municipio = r.pop("_p_municipio", None)
            ana_municipio_nome = r.pop("_ana_municipio_nome", None)
            if p_municipio and p_municipio != ana_municipio_nome:
                r["municipio_codigo_efetivo"] = mapa_ibge.get(
                    chave_nome_municipio(str(p_municipio))
                )
            estado[str(r["codigo_ana"]).strip()] = r
    return estado


def normalizar_texto(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def normalizar_numero(v) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    if isinstance(v, Decimal):
        return float(v)
    s = str(v).strip().replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def normalizar_data(v) -> date | None:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # tenta ISO
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return None


def igual(a, b, tipo: str) -> bool:
    if tipo == TIPO_TEXTO:
        return normalizar_texto(a) == normalizar_texto(b)
    if tipo == TIPO_NUMERO:
        na, nb = normalizar_numero(a), normalizar_numero(b)
        if na is None or nb is None:
            return na is None and nb is None
        return abs(na - nb) < 1e-9
    if tipo == TIPO_DATA:
        return normalizar_data(a) == normalizar_data(b)
    if tipo == TIPO_BOOL:
        return bool(a) == bool(b)
    return a == b


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--origem", required=True, help="Caminho do .xlsx baixado do SharePoint")
    ap.add_argument("--saida", default=None, help="Caminho do .xlsx de saida (default: <origem>_RESPOSTA_SPAGUAS.xlsx)")
    ap.add_argument("--dry-run", action="store_true", help="Nao salva; relata o que mudaria")
    args = ap.parse_args()

    origem = Path(args.origem).resolve()
    if not origem.exists():
        raise SystemExit(f"Arquivo nao encontrado: {origem}")
    saida = Path(args.saida) if args.saida else origem.with_name(
        origem.stem + "_RESPOSTA_SPAGUAS.xlsx"
    )

    print(f"Origem: {origem}")
    print(f"Saida:  {saida}")
    print(f"Dry-run: {args.dry_run}")
    print()

    print("1/4 Carregando estado efetivo do banco...")
    estado = buscar_estado()
    print(f"  {len(estado)} estacoes SP carregadas.")
    print()

    print("2/4 Abrindo planilha original (~2 MB, ~11.7k linhas)... isso leva ~30s")
    wb = load_workbook(origem)
    if "Inventário" not in wb.sheetnames:
        raise SystemExit(f"Aba 'Inventário' nao encontrada. Abas: {wb.sheetnames}")
    ws = wb["Inventário"]
    print(f"  Aba 'Inventário': {ws.max_row} linhas, {ws.max_column} colunas")
    print()

    print("3/4 Adicionando colunas de controle ao cabecalho...")
    col_status_idx = ws.max_column + 1
    col_just_idx = ws.max_column + 2
    ws.cell(1, col_status_idx, LABEL_STATUS)
    ws.cell(1, col_just_idx, LABEL_JUSTIFICATIVA)
    print(f"  Coluna {col_status_idx}: {LABEL_STATUS}")
    print(f"  Coluna {col_just_idx}: {LABEL_JUSTIFICATIVA}")
    print()

    print("4/4 Percorrendo linhas SP, aplicando correcoes...")
    linhas_processadas = 0
    linhas_alteradas = 0
    celulas_alteradas = 0
    nao_encontradas: list[str] = []

    for r in range(2, ws.max_row + 1):
        uf = ws.cell(r, 1).value
        if not uf or str(uf).strip().upper() != "SP":
            continue
        linhas_processadas += 1

        codigo = ws.cell(r, 2).value
        chave = str(codigo).strip() if codigo is not None else ""
        rec = estado.get(chave)
        if not rec:
            nao_encontradas.append(chave)
            continue

        linha_teve_diff = False
        for col, alias, tipo, _label in COLUNAS:
            atual = ws.cell(r, col).value
            novo = rec.get(alias)
            if novo is None:
                continue
            if not igual(atual, novo, tipo):
                if tipo == TIPO_NUMERO:
                    valor_celula = normalizar_numero(novo)
                elif tipo == TIPO_DATA:
                    valor_celula = normalizar_data(novo)
                else:
                    valor_celula = normalizar_texto(novo)
                cel = ws.cell(r, col, valor_celula)
                cel.fill = AMARELO
                celulas_alteradas += 1
                linha_teve_diff = True

        # Colunas de controle ao final
        ws.cell(r, col_status_idx, rec.get("status"))
        just = rec.get("resposta_justificativa")
        if just:
            ws.cell(r, col_just_idx, just)

        if linha_teve_diff:
            linhas_alteradas += 1

    print()
    print("=== RESUMO ===")
    print(f"Linhas SP processadas:     {linhas_processadas}")
    print(f"Linhas com alguma alteracao: {linhas_alteradas}")
    print(f"Total de celulas pintadas: {celulas_alteradas}")
    print(f"Codigos ANA nao encontrados no banco: {len(nao_encontradas)}")
    if nao_encontradas[:5]:
        print(f"  amostra: {nao_encontradas[:5]}")
    print()

    if args.dry_run:
        print("Dry-run: nao salvando.")
        return 0

    # Escrita atomica: grava em <saida>.tmp e renomeia. openpyxl nao tem
    # lock de arquivo, e duas execucoes simultaneas no mesmo destino podem
    # corromper o XLSX. O rename de mesmo volume eh atomico em Windows e
    # POSIX, garantindo que ninguem veja um arquivo a meio caminho.
    print(f"Salvando em {saida} ...")
    temp = saida.with_suffix(saida.suffix + ".tmp")
    wb.save(temp)
    temp.replace(saida)
    print("OK.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
