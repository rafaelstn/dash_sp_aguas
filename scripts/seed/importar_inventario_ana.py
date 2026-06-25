"""
Importa o inventario ANA (Meta I.6 do PROGESTAO 3) para ana_revisao_estacao.

A planilha de origem e o inventario NACIONAL: a aba 'Inventario' traz todos os
orgaos e UFs. Importamos APENAS as linhas de SP (coluna 'Responsavel - UF' == 'SP').

A planilha pode trazer (formato "Resposta SPAguas") duas colunas extras que
recuperam o trabalho de revisao de ciclos anteriores:
  STATUS_REVISAO_SPAGUAS -> status                 (enum de ana_revisao_estacao)
  JUSTIFICATIVA_SPAGUAS  -> resposta_justificativa (texto que volta no export)

Nota de schema: as colunas `correcoes` (JSONB) e `justificativa` foram dropadas
na migration 0032 (postos virou fonte unica). A migration 0039 introduziu
`resposta_*`, e `resposta_justificativa` e documentada como o campo que sai na
coluna JUSTIFICATIVA_SPAGUAS do export. Logo o re-import e o ciclo de ida e
volta desse campo. observacao_1..5 nao existem nesta planilha -> ficam NULL.

Mapeamento por NOME de coluna normalizado (lower, sem acento, sem pontuacao),
lendo o header da aba. NAO depende de indice fixo: resiste a reordenacao de
colunas em ciclos futuros.

Pipeline:
  1. Cria lote em ana_revisao_lote (hash do arquivo para idempotencia).
  2. Le a aba 'Inventario', filtra UF == 'SP', popula ana_revisao_estacao.
  3. Aplica status/resposta_justificativa importados; status != 'pendente'
     marca revisado_em; justificativa presente marca resposta_fonte.
  4. Cruza com a tabela postos (codigo_ana, depois codigo_adicional, resto sem_match).
  5. Roda bulk_analisar_divergencias() (divergencia geografica PostGIS).
  6. Registra audit em ana_revisao_evento (criada; revisada/promovida_a_posto
     quando o status importado for esses; justificada quando ha justificativa).

Idempotente: re-rodar com o mesmo arquivo (mesmo hash) substitui as estacoes do lote.
Transacional: tudo num unico commit; qualquer erro faz rollback.

Uso:
    ops/indexer/.venv/Scripts/python.exe scripts/seed/importar_inventario_ana.py \\
        "data/inventario-ana/progestao3_meta_i6_resposta_spaguas.xlsx" \\
        --nome "PROGESTAO 3 Meta I.6 - Resposta SPAguas" \\
        --prazo 2026-07-20
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
import psycopg


# Nome da aba de dados (sem acento na comparacao; ver achar_aba).
ABA_INVENTARIO = "Inventario"

# UF que importamos (a planilha e nacional).
UF_ALVO = "SP"

# Valores validos do CHECK de status em ana_revisao_estacao.
STATUS_VALIDOS = frozenset({
    "pendente",
    "em_revisao",
    "revisada",
    "descartada",
    "sem_match",
    "promovida_a_posto",
})

# Eventos de audit que disparam por status importado != 'pendente'.
STATUS_PARA_EVENTO = {
    "revisada": "revisada",
    "promovida_a_posto": "promovida_a_posto",
    "descartada": "descartada",
}


# -----------------------------------------------------------------------------
# Mapeamento por NOME de coluna normalizado -> campo do banco.
#
# A chave e o nome do header da planilha apos normalizar_chave(). Resistente a
# reordenacao de colunas: o indice fisico e descoberto na hora, lendo o header.
# As colunas 1..37 sao o inventario padrao ANA; 38..39 sao a resposta SPAguas.
# -----------------------------------------------------------------------------
MAPA_COLUNAS = {
    "responsavel uf": "responsavel_uf",
    "estacao codigo": "codigo_ana",
    "estacao nome": "nome",
    "estacao codigo adicional": "codigo_adicional",
    "latitude dec": "latitude",
    "longitude dec": "longitude",
    "latitude graus": "latitude_graus",
    "longitude graus": "longitude_graus",
    "altitude": "altitude",
    "estacao area de drenagem km2": "area_drenagem_km2",
    "baciacodigo": "bacia_codigo",
    "bacia nome": "bacia_nome",
    "subbaciacodigo": "subbacia_codigo",
    "subbacia nome": "subbacia_nome",
    "riocodigo": "rio_codigo",
    "rionome": "rio_nome",
    "estadocodigo": "estado_codigo",
    "estado sigla": "estado_sigla",
    "municipiocodigo": "municipio_codigo",
    "municipio nome": "municipio_nome",
    "responsavelcodigo": "responsavel_codigo",
    "responsavel nome": "responsavel_nome",
    "responsavel sigla": "responsavel_sigla",
    "estacao tipo": "estacao_tipo",
    "escala inicio": "escala_inicio",
    "escala fim": "escala_fim",
    "descarga liquida inicio": "descarga_liquida_inicio",
    "descarga liquida fim": "descarga_liquida_fim",
    "sedimentos inicio": "sedimentos_inicio",
    "sedimentos fim": "sedimentos_fim",
    "qualidade de agua inicio": "qualidade_inicio",
    "qualidade de agua fim": "qualidade_fim",
    "pluviometro inicio": "pluviometro_inicio",
    "pluviometro fim": "pluviometro_fim",
    "telemetria inicio": "telemetria_inicio",
    "telemetria fim": "telemetria_fim",
    "operando": "operando",
    "status revisao spaguas": "status",
    "justificativa spaguas": "resposta_justificativa",
}

# Campos que ficam fora da tabela (so usados para filtro de UF ou descartados).
CAMPOS_AUXILIARES = frozenset({
    "responsavel_uf",
    "responsavel_codigo",
    "responsavel_nome",
    "latitude_graus",
    "longitude_graus",
    "estado_codigo",
})

# Colunas da tabela ana_revisao_estacao preenchidas no INSERT (ordem estavel).
COLUNAS_INSERT = [
    "codigo_ana", "codigo_adicional", "nome",
    "latitude", "longitude", "altitude", "area_drenagem_km2",
    "bacia_codigo", "bacia_nome", "subbacia_codigo", "subbacia_nome",
    "rio_codigo", "rio_nome", "estado_sigla",
    "municipio_codigo", "municipio_nome", "responsavel_sigla", "estacao_tipo",
    "escala_inicio", "escala_fim",
    "descarga_liquida_inicio", "descarga_liquida_fim",
    "sedimentos_inicio", "sedimentos_fim",
    "qualidade_inicio", "qualidade_fim",
    "pluviometro_inicio", "pluviometro_fim",
    "telemetria_inicio", "telemetria_fim",
    "operando",
    "status", "resposta_justificativa", "resposta_fonte", "revisado_em",
]

CAMPOS_DATA = (
    "escala_inicio", "escala_fim",
    "descarga_liquida_inicio", "descarga_liquida_fim",
    "sedimentos_inicio", "sedimentos_fim",
    "qualidade_inicio", "qualidade_fim",
    "pluviometro_inicio", "pluviometro_fim",
    "telemetria_inicio", "telemetria_fim",
)

CAMPOS_NUM = ("latitude", "longitude", "altitude", "area_drenagem_km2")

CAMPOS_CODIGO = (
    "bacia_codigo", "subbacia_codigo", "rio_codigo",
    "municipio_codigo",
)

CAMPOS_STR = (
    "nome", "codigo_adicional", "bacia_nome", "subbacia_nome", "rio_nome",
    "estado_sigla", "municipio_nome", "responsavel_sigla", "estacao_tipo",
    "resposta_justificativa",
)


# -----------------------------------------------------------------------------
# Normalizacao e coercao
# -----------------------------------------------------------------------------
def normalizar_chave(s: Any) -> str:
    """lower + sem acento + so alfanumerico separado por espaco unico."""
    if s is None:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = "".join(c if c.isalnum() else " " for c in s)
    return " ".join(s.split())


def coerce_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def coerce_int_codigo(v: Any) -> Optional[str]:
    """Codigos IBGE/ANA chegam como float (1949001.0). Normaliza para string."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def coerce_num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def coerce_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def coerce_operando(v: Any) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("sim", "s", "true", "1", "yes"):
        return True
    if s in ("nao", "n", "false", "0", "no"):
        return False
    # 'não' acentuado cai aqui apos lower; trata sem depender de acento
    if normalizar_chave(s) == "nao":
        return False
    return None


# -----------------------------------------------------------------------------
# Carga de ambiente e arquivo
# -----------------------------------------------------------------------------
def carregar_database_url() -> str:
    env = Path(".env.local").read_text(encoding="utf-8")
    match = re.search(r"^DATABASE_URL\s*=\s*(.+)$", env, re.MULTILINE)
    if not match:
        raise SystemExit("DATABASE_URL ausente em .env.local")
    return match.group(1).strip().strip('"').strip("'")


def hash_arquivo(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for bloco in iter(lambda: f.read(1 << 16), b""):
            h.update(bloco)
    return h.hexdigest()


def achar_aba(wb: openpyxl.Workbook, alvo: str) -> str:
    """Acha a aba por nome normalizado (a planilha vem com acento no titulo)."""
    alvo_norm = normalizar_chave(alvo)
    for nome in wb.sheetnames:
        if normalizar_chave(nome) == alvo_norm:
            return nome
    raise SystemExit(
        f"aba '{alvo}' nao encontrada. Abas disponiveis: {wb.sheetnames}"
    )


def montar_indice_colunas(header: tuple) -> dict[str, int]:
    """
    Mapeia campo_do_banco -> indice 0-based, casando o header normalizado
    contra MAPA_COLUNAS. Colunas nao mapeadas sao ignoradas (resiliencia).
    """
    indice: dict[str, int] = {}
    for pos, titulo in enumerate(header):
        chave = normalizar_chave(titulo)
        campo = MAPA_COLUNAS.get(chave)
        if campo is not None:
            indice[campo] = pos
    return indice


def validar_indice(indice: dict[str, int]) -> None:
    """Garante que as colunas obrigatorias para o pipeline existem."""
    obrigatorias = ("responsavel_uf", "codigo_ana")
    faltando = [c for c in obrigatorias if c not in indice]
    if faltando:
        raise SystemExit(
            "colunas obrigatorias ausentes no header: "
            + ", ".join(faltando)
        )


# -----------------------------------------------------------------------------
# Extracao de linha
# -----------------------------------------------------------------------------
def extrair_linha(row: tuple, indice: dict[str, int], stats: dict) -> Optional[dict]:
    """
    Devolve dict pronto pro INSERT, ou None se a linha deve ser pulada
    (UF != SP, sem codigo_ana, ou linha vazia).
    """
    def bruto(campo: str) -> Any:
        pos = indice.get(campo)
        if pos is None or pos >= len(row):
            return None
        return row[pos]

    uf = coerce_str(bruto("responsavel_uf"))
    if uf != UF_ALVO:
        return None

    codigo_ana = coerce_int_codigo(bruto("codigo_ana"))
    if not codigo_ana:
        stats["sp_sem_codigo"] = stats.get("sp_sem_codigo", 0) + 1
        return None

    out: dict[str, Any] = {"codigo_ana": codigo_ana}

    for campo in CAMPOS_STR:
        out[campo] = coerce_str(bruto(campo))
    for campo in CAMPOS_NUM:
        out[campo] = coerce_num(bruto(campo))
    for campo in CAMPOS_CODIGO:
        out[campo] = coerce_int_codigo(bruto(campo))
    for campo in CAMPOS_DATA:
        out[campo] = coerce_date(bruto(campo))
    out["operando"] = coerce_operando(bruto("operando"))

    # Status: normaliza e valida contra o enum. Fora do enum cai pra 'pendente'.
    status_bruto = coerce_str(bruto("status"))
    status = "pendente"
    if status_bruto is not None:
        candidato = normalizar_chave(status_bruto).replace(" ", "_")
        if candidato in STATUS_VALIDOS:
            status = candidato
        else:
            stats.setdefault("status_invalido", {})
            stats["status_invalido"][status_bruto] = (
                stats["status_invalido"].get(status_bruto, 0) + 1
            )
    out["status"] = status
    # revisado_em so quando ja houve revisao (status != pendente).
    out["revisado_em"] = "__NOW__" if status != "pendente" else None
    # resposta_fonte: marca origem do texto que volta no export para a ANA.
    out["resposta_fonte"] = (
        "manual_aprovador" if out.get("resposta_justificativa") else None
    )

    return out


# -----------------------------------------------------------------------------
# Persistencia
# -----------------------------------------------------------------------------
def flush_batch(cur, lote_id, batch: list[dict]) -> list[tuple]:
    """
    Insere o batch e devolve [(id, status, tem_justificativa), ...] das
    estacoes inseridas (para gerar o audit posterior).
    """
    cols_sql = ", ".join(COLUNAS_INSERT)
    # revisado_em tem sentinela __NOW__: troca por NOW() no SQL por linha.
    resultados: list[tuple] = []
    for d in batch:
        valores: list[Any] = [lote_id]
        ph: list[str] = ["%s"]
        for c in COLUNAS_INSERT:
            v = d.get(c)
            if c == "revisado_em" and v == "__NOW__":
                ph.append("NOW()")
            else:
                ph.append("%s")
                valores.append(v)
        sql = (
            f"INSERT INTO ana_revisao_estacao (lote_id, {cols_sql}) "
            f"VALUES ({', '.join(ph)}) "
            f"RETURNING id, status, (resposta_justificativa IS NOT NULL)"
        )
        cur.execute(sql, valores)
        resultados.append(cur.fetchone())
    return resultados


def montar_eventos(resultados: list[tuple]) -> list[tuple]:
    """
    A partir do RETURNING (id, status, tem_justificativa), gera a lista de
    eventos de audit: 'criada' sempre; status -> evento correspondente;
    'justificada' quando ha resposta_justificativa.
    """
    eventos: list[tuple] = []
    for estacao_id, status, tem_justificativa in resultados:
        eventos.append((estacao_id, "criada", None))
        ev = STATUS_PARA_EVENTO.get(status)
        if ev:
            eventos.append((
                estacao_id, ev,
                "importado da resposta SPAguas (ciclo anterior)",
            ))
        if tem_justificativa:
            eventos.append((
                estacao_id, "justificada",
                "justificativa recuperada da resposta SPAguas (ciclo anterior)",
            ))
    return eventos


def registrar_audit(cur, eventos: list[tuple]) -> None:
    """eventos: [(estacao_id, evento, observacao), ...]."""
    for estacao_id, evento, observacao in eventos:
        cur.execute(
            """
            INSERT INTO ana_revisao_evento (estacao_id, evento, observacao)
            VALUES (%s, %s, %s)
            """,
            (estacao_id, evento, observacao),
        )


# -----------------------------------------------------------------------------
# Main
# -----------------------------------------------------------------------------
def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="caminho do arquivo .xlsx do inventario ANA")
    ap.add_argument("--prazo", help="data prazo de resposta (YYYY-MM-DD)", default=None)
    ap.add_argument("--nome", help="nome do lote", default="Inventario ANA")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        raise SystemExit(f"arquivo nao encontrado: {path}")

    prazo: Optional[date] = None
    if args.prazo:
        prazo = datetime.strptime(args.prazo, "%Y-%m-%d").date()

    print("=== Importando inventario ANA (filtro UF=SP) ===")
    print(f"  arquivo: {path}")
    print(f"  prazo:   {prazo}")
    print(f"  nome:    {args.nome}")
    print()

    print("  calculando hash...")
    h = hash_arquivo(path)
    print(f"  hash:    {h[:16]}...")
    print()

    print("  abrindo xlsx...")
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    aba = achar_aba(wb, ABA_INVENTARIO)
    ws = wb[aba]
    print(f"  aba: '{aba}'")

    linhas = ws.iter_rows(min_row=1, values_only=True)
    header = next(linhas)
    indice = montar_indice_colunas(header)
    validar_indice(indice)
    tem_status = "status" in indice
    tem_justificativa = "resposta_justificativa" in indice
    print(f"  colunas mapeadas: {len(indice)}/{len(header)}")
    print(f"  STATUS_REVISAO_SPAGUAS presente: {tem_status}")
    print(f"  JUSTIFICATIVA_SPAGUAS presente:  {tem_justificativa}")
    print()

    url = carregar_database_url()
    stats: dict = {}

    # prepare_threshold=None: compativel com Supabase pooler (transaction mode).
    with psycopg.connect(url, prepare_threshold=None) as conn, conn.cursor() as cur:
        # 1. Lote idempotente por hash.
        cur.execute(
            "SELECT id FROM ana_revisao_lote WHERE hash_sha256 = %s",
            (h,),
        )
        existente = cur.fetchone()
        if existente:
            lote_id = existente[0]
            print(f"  lote ja existe (id={lote_id}), substituindo estacoes...")
            cur.execute("DELETE FROM ana_revisao_estacao WHERE lote_id = %s", (lote_id,))
            cur.execute(
                "UPDATE ana_revisao_lote SET nome = %s, arquivo_origem = %s, "
                "prazo_resposta = %s WHERE id = %s",
                (args.nome, path.name, prazo, lote_id),
            )
        else:
            cur.execute(
                """
                INSERT INTO ana_revisao_lote
                  (nome, arquivo_origem, hash_sha256, prazo_resposta)
                VALUES (%s, %s, %s, %s)
                RETURNING id
                """,
                (args.nome, path.name, h, prazo),
            )
            lote_id = cur.fetchone()[0]
            print(f"  lote criado: id={lote_id}")
        print()

        # 2. Importacao (filtra UF=SP dentro de extrair_linha).
        print("  importando linhas de SP...")
        inseridas = 0
        com_justificativa = 0
        eventos: list[tuple] = []
        batch: list[dict] = []
        BATCH_SIZE = 200

        for row in linhas:
            if not any(row):
                continue
            d = extrair_linha(row, indice, stats)
            if d is None:
                continue
            if d.get("resposta_justificativa"):
                com_justificativa += 1
            batch.append(d)

            if len(batch) >= BATCH_SIZE:
                eventos.extend(montar_eventos(flush_batch(cur, lote_id, batch)))
                inseridas += len(batch)
                batch.clear()
                print(f"    {inseridas} importadas...", end="\r", flush=True)

        if batch:
            eventos.extend(montar_eventos(flush_batch(cur, lote_id, batch)))
            inseridas += len(batch)
        print(f"    {inseridas} estacoes SP importadas." + " " * 20)
        if stats.get("sp_sem_codigo"):
            print(f"    (ignoradas {stats['sp_sem_codigo']} linhas SP sem codigo_ana)")
        if stats.get("status_invalido"):
            print(f"    status fora do enum (-> pendente): {stats['status_invalido']}")
        print()

        # 3. Audit trail.
        print(f"  registrando audit ({len(eventos)} eventos)...")
        registrar_audit(cur, eventos)
        print()

        # 4. Totais do lote. total_pendencias = estacoes ainda pendentes.
        cur.execute(
            "SELECT COUNT(*) FROM ana_revisao_estacao "
            "WHERE lote_id = %s AND status = 'pendente'",
            (lote_id,),
        )
        total_pendentes = cur.fetchone()[0]
        cur.execute(
            "UPDATE ana_revisao_lote SET total_estacoes = %s, total_pendencias = %s "
            "WHERE id = %s",
            (inseridas, total_pendentes, lote_id),
        )

        # 5. Cruzamento com postos.
        print("  cruzando com tabela postos...")
        cur.execute(
            """
            UPDATE ana_revisao_estacao e
               SET posto_id = p.id, match_tipo = 'codigo_ana'
              FROM postos p
             WHERE e.lote_id = %s AND e.posto_id IS NULL
               AND p.prefixo_ana IS NOT NULL
               AND p.prefixo_ana = e.codigo_ana
            """,
            (lote_id,),
        )
        match_ana = cur.rowcount

        cur.execute(
            """
            UPDATE ana_revisao_estacao e
               SET posto_id = p.id, match_tipo = 'codigo_adicional'
              FROM postos p
             WHERE e.lote_id = %s AND e.posto_id IS NULL
               AND e.codigo_adicional IS NOT NULL
               AND p.prefixo = e.codigo_adicional
            """,
            (lote_id,),
        )
        match_adic = cur.rowcount

        cur.execute(
            """
            UPDATE ana_revisao_estacao
               SET match_tipo = 'sem_match'
             WHERE lote_id = %s AND posto_id IS NULL
            """,
            (lote_id,),
        )
        sem_match = cur.rowcount

        print(f"    match por codigo ANA:        {match_ana}")
        print(f"    match por codigo adicional:  {match_adic}")
        print(f"    sem match (so na planilha):  {sem_match}")
        print()

        # Nota: as colunas `correcoes`/`justificativa` foram dropadas na
        # migration 0032 (postos virou fonte unica). O antigo "Bucket A"
        # (sugestoes em correcoes JSONB) nao se aplica mais a este schema.

        print("  rodando analise geografica (PostGIS)...")
        cur.execute("SELECT bulk_analisar_divergencias(%s)", (lote_id,))
        print(f"    {cur.fetchone()[0]} estacoes analisadas")
        cur.execute(
            """
            SELECT divergencia_municipio, COUNT(*)
              FROM ana_revisao_estacao
             WHERE lote_id = %s
             GROUP BY divergencia_municipio
             ORDER BY 1
            """,
            (lote_id,),
        )
        print("    divergencia geografica:")
        for cls, qtd in cur.fetchall():
            print(f"      {str(cls) if cls is not None else '(NULL)':22} {qtd}")
        print()

        # Distribuicoes de validacao.
        cur.execute(
            "SELECT status, COUNT(*) FROM ana_revisao_estacao "
            "WHERE lote_id = %s GROUP BY status ORDER BY 1",
            (lote_id,),
        )
        dist_status = cur.fetchall()
        cur.execute(
            "SELECT match_tipo, COUNT(*) FROM ana_revisao_estacao "
            "WHERE lote_id = %s GROUP BY match_tipo ORDER BY 1",
            (lote_id,),
        )
        dist_match = cur.fetchall()

        conn.commit()

    print("  distribuicao de status:")
    for s, q in dist_status:
        print(f"    {s:22} {q}")
    print("  distribuicao de match:")
    for m, q in dist_match:
        print(f"    {str(m):22} {q}")
    print()
    print("=== OK ===")
    print(f"Lote: {lote_id}")
    print(f"Estacoes SP importadas: {inseridas}")
    print(f"Com justificativa:      {com_justificativa}")
    return 0


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nAbortado.")
        sys.exit(130)
