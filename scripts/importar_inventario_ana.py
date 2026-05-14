"""
Importa planilha de duvidas da ANA para a tabela ana_revisao_estacao.

Pipeline:
  1. Cria lote em ana_revisao_lote (hash do arquivo no nome para idempotencia).
  2. Le aba 'DUVIDAS' (2371 linhas) e popula ana_revisao_estacao.
  3. Cruza com tabela postos por codigo_ana e codigo_adicional.
  4. Aplica Bucket A automatico:
       - data fim faltante em pluviometro/escala/etc copiada de postos.operacao_fim_ano
       - codigo_adicional ausente copiado de postos.prefixo (quando match por codigo ANA)
  5. Roda bulk_analisar_divergencias() para popular divergencia geografica.

Idempotente: re-rodar com o mesmo arquivo (mesmo hash) substitui o lote.

Uso:
    ops/indexer/.venv/Scripts/python.exe scripts/importar_inventario_ana.py \\
        "C:/Users/win1064/Downloads/SP_AGUAS_Estações_Dúvidas.xlsx" \\
        --prazo 2026-05-18 \\
        --nome "PROGESTAO 3 ciclo 2026"
"""

from __future__ import annotations

import argparse
import hashlib
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

import openpyxl
import psycopg


# Mapeamento coluna_xlsx -> coluna_banco (DUVIDAS, 1-indexed)
COLS = {
    "responsavel_uf": 1,
    "codigo_ana": 2,
    "nome": 3,
    "codigo_adicional": 4,
    "latitude": 5,
    "longitude": 6,
    "latitude_graus": 7,
    "longitude_graus": 8,
    "altitude": 9,
    "area_drenagem_km2": 10,
    "bacia_codigo": 11,
    "bacia_nome": 12,
    "subbacia_codigo": 13,
    "subbacia_nome": 14,
    "rio_codigo": 15,
    "rio_nome": 16,
    "estado_codigo": 17,
    "estado_sigla": 18,
    "municipio_codigo": 19,
    "municipio_nome": 20,
    "responsavel_codigo": 21,
    "responsavel_nome": 22,
    "responsavel_sigla": 23,
    "estacao_tipo": 24,
    "escala_inicio": 25,
    "escala_fim": 26,
    "descarga_liquida_inicio": 27,
    "descarga_liquida_fim": 28,
    "sedimentos_inicio": 29,
    "sedimentos_fim": 30,
    "qualidade_inicio": 31,
    "qualidade_fim": 32,
    "pluviometro_inicio": 33,
    "pluviometro_fim": 34,
    "telemetria_inicio": 35,
    "telemetria_fim": 36,
    "operando": 37,
    "observacao_1": 38,
    "observacao_2": 39,
    "observacao_3": 40,
    "observacao_4": 41,
    "observacao_5": 42,
}


def carregar_database_url() -> str:
    env = Path(".env.local").read_text(encoding="utf-8")
    match = re.search(r"^DATABASE_URL\s*=\s*(.+)$", env, re.MULTILINE)
    if not match:
        raise SystemExit("DATABASE_URL ausente em .env.local")
    return match.group(1).strip().strip('"').strip("'")


def hash_arquivo(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for bloco in iter(lambda: f.read(1 << 16), b""):
            h.update(bloco)
    return h.hexdigest()


def coerce_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def coerce_int_codigo(v: Any) -> Optional[str]:
    """Codigos IBGE/ANA chegam como float (1949001.0). Normaliza para string."""
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s if s else None


def coerce_num(v: Any) -> Optional[float]:
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def coerce_date(v: Any) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s:
        return None
    # Formatos comuns: DD/MM/YYYY, YYYY-MM-DD
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def coerce_operando(v: Any) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in ("sim", "s", "true", "1", "yes"):
        return True
    if s in ("não", "nao", "n", "false", "0", "no"):
        return False
    return None


def extrair_linha(row: tuple) -> dict:
    out = {}
    for k, idx in COLS.items():
        out[k] = row[idx - 1] if idx - 1 < len(row) else None

    # Coercoes
    out["codigo_ana"] = coerce_int_codigo(out["codigo_ana"])
    out["codigo_adicional"] = coerce_str(out["codigo_adicional"])
    out["nome"] = coerce_str(out["nome"])
    out["latitude"] = coerce_num(out["latitude"])
    out["longitude"] = coerce_num(out["longitude"])
    out["altitude"] = coerce_num(out["altitude"])
    out["area_drenagem_km2"] = coerce_num(out["area_drenagem_km2"])
    out["bacia_codigo"] = coerce_int_codigo(out["bacia_codigo"])
    out["bacia_nome"] = coerce_str(out["bacia_nome"])
    out["subbacia_codigo"] = coerce_int_codigo(out["subbacia_codigo"])
    out["subbacia_nome"] = coerce_str(out["subbacia_nome"])
    out["rio_codigo"] = coerce_int_codigo(out["rio_codigo"])
    out["rio_nome"] = coerce_str(out["rio_nome"])
    out["estado_sigla"] = coerce_str(out["estado_sigla"])
    out["municipio_codigo"] = coerce_int_codigo(out["municipio_codigo"])
    out["municipio_nome"] = coerce_str(out["municipio_nome"])
    out["responsavel_sigla"] = coerce_str(out["responsavel_sigla"])
    out["estacao_tipo"] = coerce_str(out["estacao_tipo"])
    for k in (
        "escala_inicio", "escala_fim",
        "descarga_liquida_inicio", "descarga_liquida_fim",
        "sedimentos_inicio", "sedimentos_fim",
        "qualidade_inicio", "qualidade_fim",
        "pluviometro_inicio", "pluviometro_fim",
        "telemetria_inicio", "telemetria_fim",
    ):
        out[k] = coerce_date(out[k])
    out["operando"] = coerce_operando(out["operando"])
    for k in ("observacao_1", "observacao_2", "observacao_3", "observacao_4", "observacao_5"):
        out[k] = coerce_str(out[k])

    # Remove campos que nao sao colunas da tabela
    out.pop("responsavel_uf", None)
    out.pop("responsavel_codigo", None)
    out.pop("responsavel_nome", None)
    out.pop("latitude_graus", None)
    out.pop("longitude_graus", None)
    out.pop("estado_codigo", None)
    return out


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx", help="caminho do arquivo .xlsx da ANA")
    ap.add_argument("--prazo", help="data prazo de resposta (YYYY-MM-DD)", default=None)
    ap.add_argument("--nome", help="nome do lote", default="Inventario ANA")
    args = ap.parse_args()

    path = Path(args.xlsx)
    if not path.exists():
        raise SystemExit(f"arquivo nao encontrado: {path}")

    prazo: Optional[date] = None
    if args.prazo:
        prazo = datetime.strptime(args.prazo, "%Y-%m-%d").date()

    print(f"=== Importando inventario ANA ===")
    print(f"  arquivo: {path}")
    print(f"  prazo:   {prazo}")
    print(f"  nome:    {args.nome}")
    print()

    print("  calculando hash...")
    h = hash_arquivo(path)
    print(f"  hash:    {h[:16]}...")
    print()

    print("  abrindo xlsx...")
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    ws = wb["DÚVIDAS"]
    total_linhas = ws.max_row - 1
    print(f"  {total_linhas} linhas na aba DUVIDAS")
    print()

    url = carregar_database_url()
    inseridas = 0
    pendencias = 0

    # prepare_threshold=None desabilita prepared statements (compativel com
    # Supabase pooler em transaction mode, que reusa conexoes entre clientes).
    with psycopg.connect(url, prepare_threshold=None) as conn, conn.cursor() as cur:
        # Lote idempotente: hash unico evita duplicacao
        cur.execute(
            "SELECT id FROM ana_revisao_lote WHERE hash_sha256 = %s",
            (h,),
        )
        existing = cur.fetchone()
        if existing:
            lote_id = existing[0]
            print(f"  lote ja existe (id={lote_id}), substituindo estacoes...")
            cur.execute("DELETE FROM ana_revisao_estacao WHERE lote_id = %s", (lote_id,))
        else:
            cur.execute(
                """
                INSERT INTO ana_revisao_lote
                  (nome, arquivo_origem, hash_sha256, prazo_resposta)
                VALUES (%s, %s, %s, %s)
                RETURNING id
                """,
                (args.nome, path.name, h, prazo),
            )
            lote_id = cur.fetchone()[0]
            print(f"  lote criado: id={lote_id}")
        print()

        print("  importando linhas...")
        batch = []
        BATCH_SIZE = 200
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            d = extrair_linha(row)
            if not d.get("codigo_ana"):
                continue
            tem_obs = any(d.get(f"observacao_{i}") for i in range(1, 6))
            if tem_obs:
                pendencias += 1
            batch.append((lote_id, d, tem_obs))

            if len(batch) >= BATCH_SIZE:
                _flush_batch(cur, batch)
                inseridas += len(batch)
                batch.clear()
                print(f"    {inseridas}/{total_linhas}...", end="\r", flush=True)

        if batch:
            _flush_batch(cur, batch)
            inseridas += len(batch)
        print(f"    {inseridas}/{total_linhas} importadas." + " " * 20)
        print()

        # Atualiza totais no lote
        cur.execute(
            """
            UPDATE ana_revisao_lote
               SET total_estacoes = %s, total_pendencias = %s
             WHERE id = %s
            """,
            (inseridas, pendencias, lote_id),
        )

        # Cruzamento com postos
        print("  cruzando com tabela postos...")
        cur.execute(
            """
            UPDATE ana_revisao_estacao e
               SET posto_id = p.id,
                   match_tipo = 'codigo_ana'
              FROM postos p
             WHERE e.lote_id = %s
               AND e.posto_id IS NULL
               AND p.prefixo_ana IS NOT NULL
               AND p.prefixo_ana = e.codigo_ana
            """,
            (lote_id,),
        )
        match_ana = cur.rowcount

        cur.execute(
            """
            UPDATE ana_revisao_estacao e
               SET posto_id = p.id,
                   match_tipo = 'codigo_adicional'
              FROM postos p
             WHERE e.lote_id = %s
               AND e.posto_id IS NULL
               AND e.codigo_adicional IS NOT NULL
               AND p.prefixo = e.codigo_adicional
            """,
            (lote_id,),
        )
        match_adic = cur.rowcount

        cur.execute(
            """
            UPDATE ana_revisao_estacao
               SET match_tipo = 'sem_match'
             WHERE lote_id = %s AND posto_id IS NULL
            """,
            (lote_id,),
        )
        sem_match = cur.rowcount

        print(f"    match por codigo ANA:        {match_ana}")
        print(f"    match por codigo adicional:  {match_adic}")
        print(f"    sem match (so na planilha):  {sem_match}")
        print()

        # Bucket A: auto-preenchimento
        print("  bucket A: preenchendo data fim de pluviometro (op=Nao)...")
        cur.execute(
            """
            UPDATE ana_revisao_estacao e
               SET correcoes = jsonb_set(
                     COALESCE(e.correcoes, '{}'::jsonb),
                     '{pluviometro_fim_sugerido}',
                     to_jsonb(make_date(p.operacao_fim_ano, 12, 31)::text)
                   )
              FROM postos p
             WHERE e.lote_id = %s
               AND e.posto_id = p.id
               AND e.operando = FALSE
               AND e.pluviometro_inicio IS NOT NULL
               AND e.pluviometro_fim IS NULL
               AND p.operacao_fim_ano IS NOT NULL
               AND p.operacao_fim_ano > 0
            """,
            (lote_id,),
        )
        bucket_a1 = cur.rowcount

        print(f"    pluviometro_fim sugerido: {bucket_a1}")

        # Bucket A: codigo adicional ausente, preenche de postos.prefixo
        cur.execute(
            """
            UPDATE ana_revisao_estacao e
               SET correcoes = jsonb_set(
                     COALESCE(e.correcoes, '{}'::jsonb),
                     '{codigo_adicional_sugerido}',
                     to_jsonb(p.prefixo)
                   )
              FROM postos p
             WHERE e.lote_id = %s
               AND e.posto_id = p.id
               AND (e.codigo_adicional IS NULL OR e.codigo_adicional = '')
               AND p.prefixo IS NOT NULL
            """,
            (lote_id,),
        )
        bucket_a2 = cur.rowcount

        print(f"    codigo_adicional sugerido: {bucket_a2}")
        print()

        # Analise geografica
        print("  rodando analise geografica (PostGIS)...")
        cur.execute("SELECT bulk_analisar_divergencias(%s)", (lote_id,))
        analisadas = cur.fetchone()[0]
        print(f"    {analisadas} estacoes analisadas")

        cur.execute(
            """
            SELECT divergencia_municipio, COUNT(*)
              FROM ana_revisao_estacao
             WHERE lote_id = %s
             GROUP BY divergencia_municipio
             ORDER BY 1
            """,
            (lote_id,),
        )
        print("    distribuicao:")
        for cls, qtd in cur.fetchall():
            print(f"      {cls or '(NULL)':20} {qtd}")

        conn.commit()

    print()
    print("=== OK ===")
    print(f"Lote: {lote_id}")
    print(f"Estacoes importadas: {inseridas}")
    print(f"Pendencias (com observacao): {pendencias}")
    return 0


def _flush_batch(cur, batch: list) -> None:
    """Executa INSERT em lote."""
    cols = [k for k in COLS if k not in ("responsavel_uf", "responsavel_codigo", "responsavel_nome", "latitude_graus", "longitude_graus", "estado_codigo")]
    placeholders = ", ".join(["%s"] * (len(cols) + 1))
    sql = f"""
        INSERT INTO ana_revisao_estacao (lote_id, {', '.join(cols)})
        VALUES {placeholders}
    """
    # Insert linha a linha (psycopg 3 nao tem executemany super performatico aqui;
    # 2400 linhas tudo em batch eh OK pra MVP)
    for lote_id, d, _ in batch:
        valores = [lote_id] + [d.get(c) for c in cols]
        cur.execute(
            f"""
            INSERT INTO ana_revisao_estacao (lote_id, {', '.join(cols)})
            VALUES ({', '.join(['%s'] * (len(cols) + 1))})
            """,
            valores,
        )


if __name__ == "__main__":
    try:
        sys.exit(main())
    except KeyboardInterrupt:
        print("\nAbortado.")
        sys.exit(130)
